"""
GeneMark Query
Author: Matthew Lazeroff
"""

import requests
from bs4 import BeautifulSoup
import json
import time
import os
import sys #(GRyde) Needed for current solution to creating .exe
from typing import Callable, List
from subprocess import Popen, PIPE
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, colors
import Bio.Seq
import Bio.SeqFeature
import Bio.SeqRecord
from Bio import SeqIO
from PyQt5.QtCore import QSettings
from phagecommander.Utilities import RastPy, MetagenePy, Aragorn

# Genemark Domains
FILE_DOMAIN = 'http://exon.gatech.edu/GeneMark/'
GM_DOMAIN = 'http://18.220.233.194/genemark'  # Server DNA master uses
GM_HMM_DOMAIN = 'http://exon.gatech.edu/GeneMark/gmhmmp.cgi'
GMS_DOMAIN = 'http://exon.gatech.edu/GeneMark/genemarks.cgi'
HEURISTIC_DOMAIN = 'http://exon.gatech.edu/GeneMark/heuristic_gmhmmp.cgi'
GMS2_DOMAIN = 'http://exon.gatech.edu/GeneMark/genemarks2.cgi'
GLIMMER_DOMAIN = 'http://18.220.233.194/glimmer'  # Server DNA master uses

# species
# (GRyde) *****************************************************************************
# PyInstaller creates a temp folder and stores path in _MEIPASS
# In instances where script is run from python instead of .exe, will use current directory to find species.txt
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
        
    return os.path.join(base_path, relative_path)
    
# (GRyde) **************************************************************************** Original author code
# species
#species_file = os.path.join(os.path.dirname(__file__), 'species.txt')
#with open(species_file, 'r') as file:
    #SPECIES = [specie.strip() for specie in file]
# (GRyde) ***************************************************************************** end

# (GRyde) ***************************************************************************** start
# Call resource_path to determine appropriate file path based on run environment (.exe vs python)
species_file = resource_path('species.txt')
with open(species_file, 'r') as file:
    SPECIES = [specie.strip() for specie in file]
# (GRyde) ***************************************************************************** end

# tools
TOOLS = ['gm', 'hmm', 'heuristic', 'gms', 'gms2', 'prodigal', 'glimmer', 'rast', 'metagene', 'aragorn']


class Error(Exception):
    """
    Base Error Class
    """
    pass


class GeneFile:
    """
    Class for querying GeneMark tools for a DNA sequence
    """

    class GeneFileError(Error):
        """
        Raised when no result from GeneMark is returned
        """

        def __init__(self, message):
            self.message = message

    def __init__(self, sequence_file, species, prodigalLocation=None):
        """
        Constructor
        Generates necessary parameters for post requests from DNA fasta file
        :param sequence_file:
        """
        # Load DNA Sequence into memory
        input_file_data = b''
        with open(sequence_file, 'rb') as input_file:
            current_byte = input_file.read(1)
            while (current_byte):
                input_file_data += current_byte
                current_byte = input_file.read(1)

        # full path
        self.file_path = sequence_file
        # get base file name
        self.file_name = str(os.path.basename(sequence_file).split('.')[0])

        # File creation for post requests
        self.file_info = {'file': (self.file_name, input_file_data, 'application/octet-stream')}

        # Gene species - Check if compatible type, if not, exit
        if species not in SPECIES:
            raise GeneFile.GeneFileError(
                "{} is not a compatible species type - See species.txt".format(species))
        self.species = species

        # dictionary for storing query outputs
        self.query_data = {tool: '' for tool in TOOLS}

        # store prodigal location
        self.prodigalLocation = prodigalLocation

    def glimmer_query(self):
        """
        Queries Glimmer for DNA sequence
        """

        # parameters
        payload = [('sequence', self.file_info['file'][1]),
                   ('gencode', b'11'), ('topology', b'0'),
                   ('submit', b'Run GLIMMER v3.02')]

        # headers
        headers = {'User-Agent': 'GeneQuery'}

        # perform POST of file data
        file_post = requests.post(GLIMMER_DOMAIN, data=payload, headers=headers)
        file_post.raise_for_status()
        # check for job_key in response, if not raise error
        try:
            if 'job_key' not in file_post.text:
                raise GeneFile.GeneFileError("Glimmer POST #1: Invalid response")
        except GeneFile.GeneFileError:
            raise

        # get job_key from response
        job_key = file_post.text.split('=')
        payload = [(job_key[0], job_key[1])]

        # query server for output file
        # if output file is not ready, wait 2 seconds and requery
        try:
            time.sleep(2)
            return_post = requests.post(GLIMMER_DOMAIN, data=payload, headers=headers)
            return_post.raise_for_status()
            while return_post.status_code != 200:
                time.sleep(2)
                return_post = requests.post(GLIMMER_DOMAIN, data=payload, headers=headers)
                return_post.raise_for_status()
        except requests.exceptions.HTTPError as e:
            raise GeneFile.GeneFileError(
                'Glimmer Server Error: Check for DNA for proper format or check server status')

        self.query_data['glimmer'] = return_post.content.decode('utf-8')

    def genemark_query(self):
        """
        Query to GeneMark
        """
        # parameters
        payload = [('sequence', self.file_info['file'][1]),
                   ('gencode', b'11'), ('topology', b'0'),
                   ('submit', b'Run GeneMark.hmm')]

        # headers
        headers = {'User-Agent': 'GeneQuery'}

        # perform POST of file data
        file_post = requests.post(GM_DOMAIN, data=payload, headers=headers)
        file_post.raise_for_status()
        # check for job_key in response, if not raise error
        try:
            if 'job_key' not in file_post.text:
                raise GeneFile.GeneFileError("GeneMark - Invalid Response from server")
        except GeneFile.GeneFileError:
            raise

        # get job_key from response
        job_key = file_post.text.split('=')
        payload = [(job_key[0], job_key[1])]

        # query server for output file
        # if output file is not ready, wait 2 seconds and requery
        try:
            time.sleep(2)
            return_post = requests.post(GM_DOMAIN, data=payload, headers=headers)
            return_post.raise_for_status()
            # if job is not ready, HTTP response code 202 is returned
            while return_post.status_code != 200:
                time.sleep(2)
                return_post = requests.post(GM_DOMAIN, data=payload, headers=headers)
                return_post.raise_for_status()
        except requests.exceptions.HTTPError as e:
            print(e)
            raise GeneFile.GeneFileError(
                'GeneMark Server Error: Check for DNA for proper format or check server status')

        self.query_data['gm'] = return_post.content.decode('utf-8')
        # End GeneMark Lookup -------------------------------------------------------

    def genemarkhmm_query(self):
        """
        Query GeneMark Hmm
        """
        # Begin GeneMark Hmm Lookup -------------------------------------------------
        gm_hmm_data = {'sequence': '', 'org': self.species,
                       'submit': 'Start GeneMark.hmm', 'format': 'LST',
                       'subject': 'GeneMark.hmm prokaryotic',
                       'email': ''}

        # GeneMark hmm post - if unsuccessful, error thrown
        hmm_post_request = requests.post(GM_HMM_DOMAIN, files=self.file_info, data=gm_hmm_data)
        hmm_post_request.raise_for_status()
        soup = BeautifulSoup(hmm_post_request.text, 'html.parser')

        # Get URL for hmm output
        file_location = ''
        for a in soup.find_all('a', href=True):
            if 'tmp' in a['href']:
                file_location = a['href']
                break

        # if tmp not available, change in response format or invalid post
        try:
            if file_location == '':
                raise GeneFile.GeneFileError("GeneMark Hmm")
        except GeneFile.GeneFileError:
            raise

        getHmmFile = requests.get(FILE_DOMAIN + file_location)
        getHmmFile.raise_for_status()
        self.query_data['hmm'] = getHmmFile.content.decode('utf-8')
        # End GeneMark Hmm Lookup --------------------------------------------------

    def genemarks_query(self):
        """
        Query GeneMarkS
        """
        # Begin GeneMarkS Lookup ---------------------------------------------------
        gms_data = {'sequence': '', 'submit': 'Start GeneMarkS', 'mode': 'phage', 'format': 'LST',
                    'subject': 'GeneMarkS', 'gcode': 11}

        # GeneMarkS post - if unsuccessful, error thrown
        gms_post_request = requests.post(GMS_DOMAIN, files=self.file_info, data=gms_data)
        gms_post_request.raise_for_status()
        soup = BeautifulSoup(gms_post_request.text, 'html.parser')

        # Get URL for hmm output
        file_location = ''
        for a in soup.find_all('a', href=True):
            if 'tmp' in a['href']:
                file_location = a['href']
                break

        # if tmp not available, change in response format or invalid post
        try:
            if file_location == '':
                raise GeneFile.GeneFileError("GeneMarkS")
        except GeneFile.GeneFileError:
            raise

        getGmsFile = requests.get(FILE_DOMAIN + file_location)
        getGmsFile.raise_for_status()
        self.query_data['gms'] = getGmsFile.content.decode('utf-8')
        # End GeneMarkS Lookup -----------------------------------------------------

    def genemark_heuristic_query(self):
        """
        Query GeneMark Heuristic
        :return: name of file created
        """
        # Begin GeneMark Heuristic Lookup ------------------------------------------
        heuristic_data = {'sequence': '', 'submit': 'Start GeneMark.hmm', 'format': 'LST',
                          'email': '',
                          'subject': 'GeneMark.hmm', 'gcode': 11, 'strand': 'both',
                          'mod_type': 1999}

        # GeneMark Heuristic post - if unsuccessful, error thrown
        heuristic_post_request = requests.post(HEURISTIC_DOMAIN, files=self.file_info,
                                               data=heuristic_data)
        heuristic_post_request.raise_for_status()
        soup = BeautifulSoup(heuristic_post_request.text, 'html.parser')

        # Get URL for heuristic output
        file_location = ''
        for a in soup.find_all('a', href=True):
            if 'tmp' in a['href']:
                file_location = a['href']
                break

        # if tmp not available, change in response format or invalid post
        try:
            if file_location == '':
                raise GeneFile.GeneFileError("GeneMark Heuristic")
        except GeneFile.GeneFileError:
            raise

        getHeuristicFile = requests.get(FILE_DOMAIN + file_location)
        getHeuristicFile.raise_for_status()
        self.query_data['heuristic'] = getHeuristicFile.content.decode('utf-8')
        # End GeneMark Heuristic Lookup -------------------------------------------

    def genemarks2_query(self):
        """
        Query GeneMarkS2
        """
        # Begin GeneMarkS2 Lookup -------------------------------------------------
        gmms2_data = {'sequence': '', 'submit': 'GeneMarkS-2', 'mode': 'auto', 'format': 'lst',
                      'email': '', 'subject': 'GeneMarkS-2', 'gcode': 11}

        # GeneMarkS2 Post Request
        gmms2_post_request = requests.post(GMS2_DOMAIN, files=self.file_info, data=gmms2_data)
        gmms2_post_request.raise_for_status()
        soup = BeautifulSoup(gmms2_post_request.text, 'html.parser')

        # Get URL for GMS2 output
        file_location = ''
        for a in soup.find_all('a', href=True):
            if 'tmp' in a['href']:
                file_location = a['href']
                break

        # if tmp not available, change in response format or invalid post
        try:
            if file_location == '':
                raise GeneFile.GeneFileError("GeneMarkS2")
        except GeneFile.GeneFileError:
            raise

        getGMS2File = requests.get(FILE_DOMAIN + file_location)
        getGMS2File.raise_for_status()
        self.query_data['gms2'] = getGMS2File.content.decode('utf-8')
        # End GeneMarkS2 Lookup --------------------------------------------------

    def prodigal_query(self):
        """
        Calls prodigal to analyze file
        """
        # # get path for prodigal exe

        # generate prodigal command and run
        cmd = '\"{}\" -i \"{}\" -p meta'.format(self.prodigalLocation, self.file_path)
        proc = Popen(cmd, stdout=PIPE, stderr=PIPE, shell=True, stdin=subprocess.DEVNULL)
        stdout, stderr = proc.communicate()

        # check for error, exit if so
        if proc.returncode != 0:
            print(stderr)
            raise GeneFile.GeneFileError("Prodigal")

        self.query_data['prodigal'] = stdout.decode('utf-8')

    def rastQuery(self, username, password, jobId: int = None):
        """
        Submit the fasta file to RAST servers for submission
        :param username: RAST username
        :param password: RAST password
        :param jobId: RAST jobID
        :return:
        """
        # create RAST object
        rastJob = RastPy.Rast(username, password, jobId=jobId)

        # if a jobID was given, check if it is complete
        if jobId is None or not rastJob.checkIfComplete():
            # submit
            rastJob.submit(self.file_path, self.file_name)

            # check periodically for job completion
            RAST_COMPLETION_CHECK_DELAY = 15
            time.sleep(RAST_COMPLETION_CHECK_DELAY)
            while not rastJob.checkIfComplete():
                time.sleep(RAST_COMPLETION_CHECK_DELAY)

        # job is complete - retrieve gene annotation
        self.query_data['rast'] = rastJob.retrieveData()

    def metageneQuery(self):
        """
        Query Metagene servers for analysis
        """
        metaGene = MetagenePy.Metagene(self.file_path, self.file_name)
        self.query_data['metagene'] = metaGene.query()

    def aragornQuery(self):
        self.query_data['aragorn'] = Aragorn.aragorn_query(self.file_path)


class GeneError(Error):
    def __init__(self, message):
        self.message = message


class GeneFeature:
    DIRECTIONS = {'+', '-'}

    # (GRyde) Updated with totalLength parameter
    def __init__(self, start: int, stop: int, direction: str, totalLength):
        self.start = start
        self.stop = stop

        if direction not in GeneFeature.DIRECTIONS:
            raise TypeError(f'Invalid direction {direction}')
        self.direction = direction

        self.length = stop - start + 1
        # (GRyde) Length calculation for genes that wrap around back to their start
        if self.length < 0:
            self.length = totalLength + stop - start + 1

    def __eq__(self, other):

        if not isinstance(other, GeneFeature):
            raise TypeError(f'== cannot be used with type: {type(other)}')
        else:
            if self.direction == other.direction:
                if self.direction == '+':
                    if self.stop == other.stop:
                        return True
                else:  # - direction
                    if self.start == other.start:
                        return True

        return False


class Gene(GeneFeature):
    """
    Class for representing a potential gene encoding
    """

    # (GRyde) Updated with totalLength parameters
    def __init__(self, start: str, stop: str, direction: str, identity='', totalLength=0):
        """
        Constructor
        :param start:  start codon (str)
        :param stop:   end codon (str)
        :param direction: +/-
        :param identity: optional identifier
        """
        # check for "<3" or ">3" style starts, stops
        if '<' in start:
            start = start.split('<')[-1]
            start = int(start)
        elif '&lt;' in start:
            start = start.split('&lt;')[-1]
            start = int(start)
        else:
            start = int(start)

        if '>' in stop:
            stop = stop.split('>')[-1]
            stop = int(stop)
        elif '&gt;' in stop:
            stop = stop.split('&gt;')[-1]
            stop = int(stop)
        else:
            stop = int(stop)

        self.identity = identity

        # (GRyde) Calls GeneFeature __init__, I added totalLength as before it was:
        # super(Gene, self).__init__(start, stop, direction)
        super(Gene, self).__init__(start, stop, direction, totalLength)

    def jsonDump(self):

        data = {'start': self.start,
                'stop': self.stop,
                'direction': self.direction,
                'length': self.length,
                'identity': self.identity}

        return data

    # def __eq__(self, other):
    #     """
    #     Checks if Genes can possibly represent the same Gene
    #     For + direction, True is stop codons are same
    #     For - direction, True if start codons are same
    #     :param other: Gene object
    #     :return: True / False
    #     """
    #     try:
    #         if isinstance(other, GeneFeature):
    #             if self.direction == other.direction:
    #                 if self.direction == '+':
    #                     if self.stop == other.stop:
    #                         return True
    #                 else:  # - direction
    #                     if self.start == other.start:
    #                         return True
    #         else:
    #             raise GeneError("Gene Eq: Comparing object must be of Gene type")
    #     except GeneError:
    #         raise
    #
    #     return False

    def __str__(self):
        """
        Overload of str() for printing
        :return: String containing gene information
        """
        return 'Direction: {} Start: {} Stop: {} Length: {}'.format(self.direction,
                                                                    self.start,
                                                                    self.stop,
                                                                    self.length)

    def __repr__(self):
        return '({}, {}, {})'.format(self.direction, self.start, self.stop)


"""
    (GRyde)
    Original TRNA __init__ method:
    def __init__(self, start, stop, direction, trna_type, identity=None):
        super(TRNA, self).__init__(start, stop, direction)
        
    Updated to include totalLength as parameter and use totalLength in inheritance statement, super(), to GeneFeature
"""
class TRNA(GeneFeature):

    def __init__(self, start, stop, direction, trna_type, totalLength, identity=None):
        super(TRNA, self).__init__(start, stop, direction, totalLength)

        self.type = trna_type
        self.identity = identity

    def __repr__(self):
        return f'TRNA(start={self.start}, stop={self.stop}, direction={self.direction}, type={self.type}'


class GeneUtils:
    """
    Class for operations relating to Genes
    """

    @staticmethod
    def getGeneComparison(gene: Gene) -> int:
        """
        Helper function to retrieve the stop/start of a gene depending on its direction
        :param gene: Gene
        :return: the start/stop of a Gene
        """
        if gene.direction == '+':
            return gene.stop
        else:
            return gene.start

    @staticmethod
    def sortGenes(genes: List[Gene]) -> List[Gene]:
        """
        Sort Genes according to their start/stop depending on the direction of the gene
        * Forward direction genes are sorted by their stop
        * Negative direction genes are sorted by their start
        :param genes: List of Genes to sort
        :return: List[Gene] sorted by start/stop depending on direction of gene
        """
        return sorted(genes, key=GeneUtils.getGeneComparison)

    # (GRyde) Adding third parameter to this fuction to allowing for tRNA to be exported
    #         regardless of spinbox value
    # Originally: (genes: List[Gene], comparisonFunc: Callable[[int], bool]) -> List[List[Gene]]
    @staticmethod
    def filterGenes(genes: List[Gene], comparisonFunc: Callable[[int], bool], exportRNA: bool) -> List[List[Gene]]:
        """
        Filters the genes to only those where there are <limit> or more of that gene
        Ex: Filter for genes which there are more than 3 of each
            greaterThanThreeGenes = filterGenes(genes, lambda x: x > 3)

        :param genes: List[Gene]
        :param comparisonFunc: a function which takes a quantity and returns a bool based on that value
            * Arg 1: Quantity (int)
            * Result: bool
            * Ex: lambda x: x <= 10
        :return: List[List[Gene]] in order of stop/starts
        """
        filteredGenes = []
        sortedGenes = GeneUtils.sortGenes(genes)

        # group the genes according to their stops/starts
        # discard groups with less than <limit> items
        currentGroup = [sortedGenes[0]]
        previousGene = sortedGenes[0]
        for gene in sortedGenes[1:]:
            # if the current gene is the same as the previous, add to the same group
            if gene == previousGene:
                currentGroup.append(gene)

            # different genes, create a new group
            else:
                # if comparison is satisfactory, add to genes to be returned
                # else, they're dropped
                
                # (GRyde) If exporting tRNA, check group against filter amount.
                #         If doesn't match, check if TRNA obj first before discarding.
                if exportRNA:
                    if comparisonFunc(len(currentGroup)):
                        filteredGenes.append(currentGroup)
                    
                    elif isinstance(currentGroup[0], TRNA):
                        filteredGenes.append(currentGroup)
                # (GRyde) If not exporting tRNA, check first if group is made of TRNA obj(s).
                #         If yes, no action and move to discard statement.
                #         If no, check group against selected filter function and append or not as normal.
                else:
                    if isinstance(currentGroup[0], TRNA):
                        pass # Pass since user does NOT want TRNA
                    elif comparisonFunc(len(currentGroup)):
                        filteredGenes.append(currentGroup)
                    
                # new group of genes
                currentGroup = [gene]

            previousGene = gene

        # add last group if sufficient
        if comparisonFunc(len(currentGroup)):
            filteredGenes.append(currentGroup)

        return filteredGenes

    @staticmethod
    def genbankToFile(sequence: str, genes: List[Gene], fileName: str):
        """
        Writes the list of Genes to file in genbank format
        :param sequence: DNA sequence
        :param genes: list of Genes
        :param fileName: name of the file to write to
        """
        # create sequence from sequence string
        seq = Bio.Seq.Seq(sequence)

        # sort genes from smallest to largest starts
        genes = GeneUtils.sortGenes(genes)

        # build features
        features = []
        for ind, gene in enumerate(genes):
            ind += 1
            direction = 1 if gene.direction == '+' else -1
            # (GRyde) ****************************************************************** start
            if gene.start > gene.stop:
                firstJoinLocation = Bio.SeqFeature.FeatureLocation(gene.start - 1, len(seq))
                secondJoinLocation = Bio.SeqFeature.FeatureLocation(0, gene.stop)
                combinedLocation = Bio.SeqFeature.CompoundLocation([firstJoinLocation, secondJoinLocation])
                
                geneFeature = Bio.SeqFeature.SeqFeature(combinedLocation,
                                                        type='gene',
                                                        qualifiers={'gene': ind},
                                                        strand=direction)
                                                        
                if isinstance(gene, Gene):
                    cdsFeature = Bio.SeqFeature.SeqFeature(combinedLocation,
                                                           type='CDS',
                                                           qualifiers={'gene': ind},
                                                           strand=direction)
                elif isinstance(gene, TRNA):
                    product = gene.type.split('(')[0]
                    cdsFeature = Bio.SeqFeature.SeqFeature(combinedLocation,
                                                           type='TRNA',
                                                           qualifiers={'gene': ind,
                                                                       'note': gene.type,
                                                                       'product': product},
                                                           strand=direction)
            else:
                geneFeature = Bio.SeqFeature.SeqFeature(Bio.SeqFeature.FeatureLocation(gene.start - 1, gene.stop),
                                                        type='gene',
                                                        qualifiers={'gene': ind},
                                                        strand=direction)
                                                        
                if isinstance(gene, Gene):
                    cdsFeature = Bio.SeqFeature.SeqFeature(Bio.SeqFeature.FeatureLocation(gene.start - 1, gene.stop),
                                                           type='CDS',
                                                           qualifiers={'gene': ind},
                                                           strand=direction)
                elif isinstance(gene, TRNA):
                    product = gene.type.split('(')[0]
                    cdsFeature = Bio.SeqFeature.SeqFeature(Bio.SeqFeature.FeatureLocation(gene.start - 1, gene.stop),
                                                           type='TRNA',
                                                           qualifiers={'gene': ind,
                                                                       'note': gene.type,
                                                                       'product': product},
                                                           strand=direction)
            # (GRyde) Original code below in case catastrophic failure
            # geneFeature = Bio.SeqFeature.SeqFeature(Bio.SeqFeature.FeatureLocation(gene.start - 1, gene.stop),
                                                    # type='gene',
                                                    # qualifiers={'gene': ind},
                                                    # strand=direction)
            # if isinstance(gene, Gene):
                # cdsFeature = Bio.SeqFeature.SeqFeature(Bio.SeqFeature.FeatureLocation(gene.start - 1, gene.stop),
                                                       # type='CDS',
                                                       # qualifiers={'gene': ind},
                                                       # strand=direction)
            # elif isinstance(gene, TRNA):
                # product = gene.type.split('(')[0]
                # cdsFeature = Bio.SeqFeature.SeqFeature(Bio.SeqFeature.FeatureLocation(gene.start - 1, gene.stop),
                                                       # type='TRNA',
                                                       # qualifiers={'gene': ind,
                                                                   # 'note': gene.type,
                                                                   # 'product': product},
                                                       # strand=direction)
            # (GRyde) ****************************************************************** end
            features.append(geneFeature)
            features.append(cdsFeature)

        # create genbank file from genes and write to file
        """
            (GRyde) Code before adding molecule_type annotation
            gbRecord = Bio.SeqRecord.SeqRecord(seq, features=features,
                                           name=os.path.split(fileName)[1].split('.')[0])
        """
        # (GRyde) As of BioPython version 1.78, molecule_type needs to be explicitly stated in order to write/export
        # to a .gb (GenBank) file
        
        # (GRyde) The GitHub for the BioPython SeqIO class shows that this exact check (len(locus.split()) > 1) is used to determine if whitespace error should be raised.
        #         So basically borrowing that check to anticipate and correct/prevent errors
        #         - File name will still keep the intended spaces used. Only replaces whitespace with '_'s in header of .gb file
        locusLineName = os.path.split(fileName)[1].split('.')[0]
        if len(locusLineName.split()) > 1:
            validLocusName = locusLineName.replace(' ', '_')
        else:
            validLocusName = locusLineName
            
        gbRecord = Bio.SeqRecord.SeqRecord(seq, features=features,
                                           name=validLocusName,
                                           annotations={'molecule_type': 'DNA'})
        
        """
        (GRyde) Original approach to creating gbRecord obj. Need to change so that '_' can be added to names with spaces since .gb files do NOT like titles with spaces
                though it is ok for the file name to have spaces
        gbRecord = Bio.SeqRecord.SeqRecord(seq, features=features,
                                           name=os.path.split(fileName)[1].split('.')[0],
                                           annotations={'molecule_type': 'DNA'})
        """
        SeqIO.write([gbRecord], fileName, 'genbank')

    @staticmethod
    def findMostGeneOccurrences(genes: List[Gene]) -> Gene:
        """
        Takes a list of the same Gene and returns the Gene with the most occurrences.
        If there are equal occurrences of the Gene, the longest one will be chosen
        :param genes: list of Genes
            * Genes should be the same
                - If positive, same stops
                - If negative, same starts
                - Error is thrown if all the Genes are not the same
        :return: Gene
        """
        # dictionary: repr(Gene): [<Gene>, occurrences]
        geneOccurrences = dict()
        # calculate frequencies of each gene
        currGene = genes[0]
        for gene in genes:
            # check if not the same gene, if so, throw error
            if gene != currGene:
                raise ValueError('{} does not match {}'.format(repr(gene), repr(currGene)))
            currGene = gene

            geneStr = repr(gene)
            if geneStr in geneOccurrences:
                geneOccurrences[geneStr][1] += 1
            else:
                geneOccurrences[geneStr] = [gene, 1]

        # sort genes according to frequency
        # maxGenes = List of [<Gene>, frequency]
        maxGenes = sorted(geneOccurrences.values(), key=lambda x: x[1], reverse=True)

        maxFrequencyGene = maxGenes[0]
        if len(maxGenes) > 1:
            # check if more than one gene share the same max frequency
            for gene in maxGenes[1:]:
                # if the gene has the same frequency as the maximum, pick the longest one
                if gene[1] == maxFrequencyGene[1]:
                    # see if current gene is longer than the current max
                    if gene[0].length > maxFrequencyGene[0].length:
                        maxFrequencyGene = gene

        return maxFrequencyGene[0]
        
    # (GRyde) Choose genes based on longest
    #         See findMostGeneOccurrences for notes since methods are very similar
    @staticmethod
    def findLongestGene(genes: List[Gene]) -> Gene:
        """
        Takes a list of the same Gene and returns the longest one.
        Shouldn't be possible to have ties for longest
        """
        
        currGene = genes[0]
        for gene in genes:
            if gene != currGene:
                raise ValueError('{} does not match {}'.format(repr(gene), repr(currGene)))
            currGene = gene
        
        longestGene = genes[0]
        for gene in genes:
            if gene.length > longestGene.length:
                longestGene = gene
                
        return longestGene
        
    # (GRyde) Choose genes based on specific program selection
    #         See findMostGeneOccurrences for notes since methods are fairly similar
    @staticmethod
    def useSpecificProgram(genes: List[Gene], program) -> Gene:
        """
        Takes a list of the same Gene and returns the one from the selected program.
        In the event that a gene from the selected program isn't available, the gene with
        the most occurrences is used. If there is a tie for most occurrences, the longest
        gene is used.
        """
        
        geneOccurrences = dict()
        
        currGene = genes[0]
        for gene in genes:
            if gene != currGene:
                raise ValueError('{} does not match {}'.format(repr(gene), repr(currGene)))
            currGene = gene
            
            geneStr = repr(gene)
            if geneStr in geneOccurrences:
                geneOccurrences[geneStr][1] += 1
            else:
                geneOccurrences[geneStr] = [gene, 1]
                
        for gene in genes:
            if gene.identity == program:
                
                return gene
                
        maxGenes = sorted(geneOccurrences.values(), key=lambda x: x[1], reverse=True)
        
        maxFrequencyGene = maxGenes[0]
        if len(maxGenes) > 1:
            for gene in maxGenes[1:]:
                if gene[1] == maxFrequencyGene[1]:
                    if gene[0].length > maxFrequencyGene[0].length:
                        maxFrequencyGene = gene
                        
        return maxFrequencyGene[0]
        


class GeneParse:
    """
    Class for parsing GeneMark output files
    All methods are static
    """
    
    """ (GRyde) 
         - All parse methods given additional parameter, totalLength
            - Prior versions only used tool_data and identity = ''
         - All instances of creating Gene obj. updated with additional argument, totalLength=totalLength
            - Prior versions ended at identity=identity
         - Would have been tons of work to re-order the parameter list so giving totalLength a default value of 0 so it can come after identity
    """
    @staticmethod
    def parse_glimmer(glimmer_data, identity='', totalLength=0):
        """
        Parses the output of a glimmer query for gene predictions
        :param glimmer_data: string representing glimmer query output
        :param identity: optional identifier for each gene
        :return: list of Genes in numerical order
        """
        # split text into lines - skipping first line
        data = glimmer_data.splitlines()[1:]

        # Get Gene data
        genes = []
        for line in data:
            if 'html' not in line:
                data = [x for x in line.split(' ') if x != '']
                # get gene direction and create Gene
                if '+' in data[3]:
                    genes.append(Gene(data[1], data[2], '+', identity=identity, totalLength=totalLength))
                else:
                    genes.append(Gene(data[2], data[1], '-', identity=identity, totalLength=totalLength))

        # return list of Genes
        return genes

    @staticmethod
    def parse_genemark(gm_data, identity='', totalLength=0):
        """
        Parse GeneMark output file for Genes
        :param gm_file: GeneMark output file
        :param identity: optional identifier for each Gene
        :return: list of Genes in file order
        """
        gm_data = gm_data.splitlines()

        # skip until Region of Interests section
        index = 0
        for ind, line in enumerate(gm_data):
            if '   Gene    Strand    LeftEnd    RightEnd       Gene     Class' in line:
                index = ind
                break

        # invalid file format
        if index == len(gm_data) - 1:
            raise GeneError('Invalid Genemark file format')

        # go to first gene
        gm_data = gm_data[index + 2:]

        # Get gene data
        genes = []
        for line in gm_data:
            if line != '':
                data = [x for x in line.strip().split(' ') if x != '']
                start = data[2]
                stop = data[3]
                direction = data[1]
                genes.append(Gene(start, stop, direction, identity=identity, totalLength=totalLength))
            else:
                break

        return genes

    @staticmethod
    def parse_genemarkS(gms_data, identity='', totalLength=0):
        """
        Parse GeneMarkS file for Gene data
        :param s_file: GeneMarkS file
        :param identity: optional identifier for each Gene
        :return: list of Genes in file order
        """
        gms_data = gms_data.splitlines()

        # begin parse
        current_line = gms_data[0]
        # Check if proper GeneMark file
        try:
            if 'GeneMark.hmm' not in current_line:
                raise GeneError("GMS: Not a valid GeneMark GMS file")
        except GeneError:
            raise

        # skip through until gene data
        index = 0
        for ind, current_line in enumerate(gms_data[1:]):
            if 'Gene    Strand    LeftEnd    RightEnd       Gene     Class' in current_line:
                index = ind
                break

        gms_data = gms_data[index + 3:]

        # array for gene objs
        genes = []
        # get gene data
        for current_line in gms_data:
            if current_line != '':
                data = [x for x in current_line.strip().split(' ') if x != '']
                genes.append(Gene(data[2], data[3], data[1], identity=identity, totalLength=totalLength))

        return genes

    @staticmethod
    def parse_genemarkHmm(hmm_data, identity='', totalLength=0):
        """
        Parse GeneMark Hmm file for Gene data
        :param hmm_file:
        :param identity: optional identifier for each Gene
        :return: list of Genes in file order
        """
        hmm_data = hmm_data.splitlines()

        # begin parse
        current_line = hmm_data[0]
        # Check if proper GeneMark file
        try:
            if 'GeneMark.hmm' not in current_line:
                raise GeneError("Hmm: Not a valid GeneMark Hmm file")
        except GeneError:
            raise

        # skip through until gene data
        # while ('Gene    Strand    LeftEnd    RightEnd       Gene     Class' not in current_line):
        #     current_line = file.readline()
        # current_line = file.readline()  # read blank line
        index = 0
        for ind, current_line in enumerate(hmm_data[1:]):
            if 'Gene    Strand    LeftEnd    RightEnd       Gene     Class' in current_line:
                index = ind
                break
        hmm_data = hmm_data[index + 3:]

        # array for gene objs
        genes = []
        # get gene data
        for current_line in hmm_data:
            if current_line != '':
                data = [x for x in current_line.strip().split(' ') if x != '']
                # check for weird characters with gene sequence ends (Ex: start = '<2')
                nonNumChars = False
                for ind, char in enumerate(data[2]):
                    if not char.isnumeric():
                        nonNumChars = True
                    elif char.isnumeric():
                        if not nonNumChars:
                            break
                        else:
                            data[2] = data[2][ind:]

                genes.append(Gene(data[2], data[3], data[1], identity=identity, totalLength=totalLength))

        return genes

    @staticmethod
    def parse_genemarkHeuristic(heuristic_data, identity='', totalLength=0):
        """
        Parse GeneMark Heuristic file for Gene data
        :param heuristic_file: GeneMark Heuristic output file
        :param identity: optional identifier for each Gene
        :return: list of Genes in order
        """
        heuristic_data = heuristic_data.splitlines()
        # begin parse
        current_line = heuristic_data[0]
        # Check if proper GeneMark file
        try:
            if 'GeneMark.hmm' not in current_line:
                raise GeneError("Hmm: Not a valid GeneMark Hmm file")
        except GeneError:
            raise

        # skip through until gene data
        index = 0
        for ind, current_line in enumerate(heuristic_data[1:]):
            if 'Gene    Strand    LeftEnd    RightEnd       Gene     Class' in current_line:
                index = ind
                break

        heuristic_data = heuristic_data[index + 3:]

        # arary for gene objs
        genes = []
        # get gene data
        for current_line in heuristic_data:
            if current_line != '':
                data = [x for x in current_line.strip().split(' ') if x != '']
                genes.append(Gene(data[2], data[3], data[1], identity=identity, totalLength=totalLength))

        return genes

    @staticmethod
    def parse_genemarkS2(gms2_data, identity='', totalLength=0):
        """
        Parse GeneMark S2 file for Gene data
        :param s2_file: GeneMark S2 output file
        :param identity: optional identifier for each Gene
        :return: list of Genes in file order
        """
        gms2_data = gms2_data.splitlines()

        # skip until gene data
        # curr_line = file.readline()
        # while ('SequenceID' not in curr_line):
        #     curr_line = file.readline()

        index = 0
        for ind, line in enumerate(gms2_data):
            if 'SequenceID' in line:
                index = ind
                break
        gms2_data = gms2_data[index + 1:]
        # (GRyde) Print last item in list to see if error-prone data is being put into Gene obj
        print(gms2_data[-1])
        # (GRyde)

        # Get gene data
        genes = []
        for curr_line in gms2_data:
            """
                (GRyde)
                - Searching for '#' in a line instead of blank works because the output of the GMS2 program seems to have some comment lines at bottom
                of data dump before reaching just blank lines. This parse method was originally taking those comment strings and trying to use that data to 
                create Gene objects which were not compatible with strings (Gene can accept str but it inherits from GeneFeature which explicitly requries int
                data type
            """
            if '#' not in curr_line: # (GRyde) Changed
            # if curr_line != '': (GRyde) This is the original line, use this to replace if catastrophic failure
                data = [x for x in curr_line.strip().split(' ') if x != '']
                genes.append(Gene(data[2], data[3], data[1], identity=identity, totalLength=totalLength))
            # stop at newline after genes
            else:
                break

        return genes

    @staticmethod
    def parse_prodigal(prodigal_data, identity='', totalLength=0):
        """
        Parse prodigal output file for Gene information
        :param prodigal_file: prodigal output file
        :param identity: optional identifier for Gene
        :return: list of Genes in file order
        """
        prodigal_data = prodigal_data.splitlines()

        # gene info starts on third line
        genes = []
        for index, line in enumerate(prodigal_data):
            if index >= 2:
                if 'CDS' in line:
                    gene_str = line.strip().split('CDS')[-1].strip()
                    direction = '+'
                    if 'complement' in gene_str:
                        direction = '-'
                        # remove 'complement'
                        gene_str = gene_str[len('complement'):]
                        # remove ()
                        gene_str = gene_str[1:-1]
                    start, end = gene_str.split('..')
                    # start, end = [int(x) for x in gene_str.split('.') if x.isnumeric()]
                    # create gene
                    genes.append(Gene(start, end, direction, identity=identity, totalLength=totalLength))

        return genes

    @staticmethod
    def parse_rast(rast_data, identity='', totalLength=0):
        """
        Parse the gff3 formatted data for genes
        :param rast_data: gff3 formatted gene annotations
        :param identity: optional identity for genes
        :return: List[Gene]
        """
        # find non comment lines
        genes = []
        for line in rast_data.splitlines():
            if 'CDS' in line:
                data = line.split('\t')
                start = data[3]
                stop = data[4]
                direction = data[6]
                genes.append(Gene(start, stop, direction, identity, totalLength))

        return genes

    @staticmethod
    def parse_metagene(metagene_data: str, identity: str = '', totalLength=0):
        """
        Parse a metagene output file for Genes
        :param metagene_data: output of Metagene tool
        :param identity: optional identity for gene
        :return: List[Gene]
        """
        return MetagenePy.Metagene.parse(metagene_data, identity, totalLength)

    @staticmethod
    def parse_aragorn(aragorn_data: str, identity: str = '', totalLength=0):
        """
        Parse Aragorn output
        :param aragorn_data:
        :param identity:
        :return: List[TRNA]
        """
        return Aragorn.aragorn_parse(aragorn_data, totalLength, id=identity)


def write_gene(gene, row, ws, indexes):
    """
    Writes the passed gene into its corresponding columns in the Excel worksheet
    :param gene: Gene object
    :param row: row to write to
    :param ws: openpyxl worksheet object
    :param indexes: dictionary of indexes, organized by gene.identity labels
    """
    left = indexes[gene.identity][0][0] + str(row)
    right = indexes[gene.identity][0][1] + str(row)
    indexes[gene.identity][1] += 1

    # write data
    curr_row = ws[left:right][0]
    curr_row[0].value = indexes[gene.identity][1]
    curr_row[1].value = gene.direction
    curr_row[1].alignment = Alignment(horizontal='center')
    curr_row[2].value = gene.start
    curr_row[3].value = gene.stop
    curr_row[4].value = gene.length


def color_row(ws, row, color):
    """
    Colors the row according to passed color
    Changes text color to white
    :param ws: worksheet to alter
    :param row: row number
    :param color: openpyxl Fill profile
    """
    # Colors which need white font
    need_white = {PatternFill(fgColor='215967', fill_type='solid'),
                  PatternFill(fgColor='31869b', fill_type='solid'),
                  PatternFill(fgColor='92cddc', fill_type='solid')}

    row = ws['A' + str(row):'AO' + str(row)][0]
    for cell in row:
        cell.fill = color
        if color in need_white:
            cell.font = Font(color=colors.WHITE)


def excel_write(output_directory, files, sequence):
    """
    Writes the content of the Gene files to a spreadsheet
    :param output_directory: directory to output to
    :param files: list of file names returned by GeneFile.query_all()
    :param sequence: GeneFile object
    """
    # output to Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Gene Calls'

    # Annotation IDs
    ids = ['GLIMMER', 'GM', 'HMM', 'GMS', 'GMS2', 'HEURISTIC']

    indexes = dict()
    indexes['GLIMMER'] = [['A', 'E'], 0]
    indexes['GM'] = [['G', 'K'], 0]
    indexes['HMM'] = [['M', 'Q'], 0]
    indexes['GMS'] = [['S', 'W'], 0]
    indexes['GMS2'] = [['Y', 'AC'], 0]
    indexes['HEURISTIC'] = [['AE', 'AI'], 0]
    indexes['PRODIGAL'] = [['AK', 'AO'], 0]
    headers_names = ['No.', 'Direction', 'Start', 'Stop', 'Length']

    # Row colors according to # of same genes
    colors = dict()
    colors[7] = PatternFill(fgColor='215967', fill_type='solid')
    colors[6] = PatternFill(fgColor='215967', fill_type='solid')
    colors[5] = PatternFill(fgColor='31869b', fill_type='solid')
    colors[4] = PatternFill(fgColor='92cddc', fill_type='solid')
    colors[3] = PatternFill(fgColor='b7dee8', fill_type='solid')
    colors[2] = PatternFill(fgColor='daeef3', fill_type='solid')
    colors[1] = PatternFill(fgColor='ffffff', fill_type='solid')

    # Format Annotation Columns
    for x in indexes.items():
        value = x[1]
        first = ws[value[0][0] + str(1): value[0][1] + str(1)][0]
        # Gene Program Label
        for column in first:
            column.value = x[0]
            column.font = Font(bold=True)
            column.alignment = Alignment(horizontal='center')

        # Column Headers
        second = ws[value[0][0] + str(2): value[0][1] + str(2)][0]
        for index, column in enumerate(second):
            column.value = headers_names[index]
            column.alignment = Alignment(horizontal='center')

    # Format Statistics Columns
    stat_row = ws['AQ1': 'AS1'][0]
    stat_headers = ['SUM', 'ALL', 'ONE']
    for index, cell in enumerate(stat_row):
        cell.value = stat_headers[index]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # get genes
    genes = []
    # parse files depending on file type and add genes to genes list
    for file in files:
        file_type = os.path.split(file)[1].split('.')[1]
        if file_type == 'glimmer':
            genes += GeneParse.parse_glimmer(file, identity='GLIMMER')
        elif file_type == 'gm':
            genes += GeneParse.parse_genemark(file, identity='GM')
        elif file_type == 'gmhmm':
            genes += GeneParse.parse_genemarkHmm(file, identity='HMM')
        elif file_type == 'gms':
            genes += GeneParse.parse_genemarkS(file, identity='GMS')
        elif file_type == 'gms2':
            genes += GeneParse.parse_genemarkS2(file, identity='GMS2')
        elif file_type == 'heuristic':
            genes += GeneParse.parse_genemarkHeuristic(file, identity='HEURISTIC')
        elif file_type == 'prodigal':
            genes += GeneParse.parse_prodigal(file, identity='PRODIGAL')

    total = sorted(genes, key=lambda x: x.start)

    # write to file
    row = 3
    count = 0
    genes_in_row = 1
    while count < len(total):
        # get next gene
        curr_gene = total[count]
        # print same genes on the same line, different ones on different lines
        if count != 0:
            if curr_gene != total[count - 1]:
                # End of current row of genes
                # update row color according to number of genes
                color_row(ws, row, colors[genes_in_row])
                # update stats of row
                stat_row = ws['AQ' + str(row): 'AS' + str(row)][0]
                # SUM
                stat_row[0].value = genes_in_row
                stat_row[0].alignment = Alignment(horizontal='center')
                # ALL
                if genes_in_row == len(indexes):
                    stat_row[1].value = 'X'
                    stat_row[1].alignment = Alignment(horizontal='center')
                # ONE
                if genes_in_row == 1:
                    stat_row[2].value = 'X'
                    stat_row[2].alignment = Alignment(horizontal='center')
                # move to next row and reset genes count
                row += 1
                genes_in_row = 1
            else:
                genes_in_row += 1

        write_gene(curr_gene, row, ws, indexes)
        count += 1

    # apply color and stats for last row
    color_row(ws, row, colors[genes_in_row])
    # update stats of row
    stat_row = ws['AQ' + str(row): 'AS' + str(row)][0]
    # SUM
    stat_row[0].value = genes_in_row
    stat_row[0].alignment = Alignment(horizontal='center')
    # ALL
    if genes_in_row == len(indexes):
        stat_row[1].value = 'X'
        stat_row[1].alignment = Alignment(horizontal='center')
    # ONE
    if genes_in_row == 1:
        stat_row[2].value = 'X'
        stat_row[2].alignment = Alignment(horizontal='center')

    # save file
    wb.save(output_directory + sequence.name + '.xlsx')


if __name__ == '__main__':
    file = 'D:\mdlaz\Documents\College\Research\programs\GeneQuery\\tests\sequences\Ronan.fasta'
    for seq in SeqIO.parse(file, 'fasta'):
        Dissequence = seq
    gfile = GeneFile(file, 'Paenibacillus_larvae_subsp_ATCC_9545')
    data = gfile.query_data['rast']
