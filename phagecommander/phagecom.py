import os
import pickle
import pathlib
from typing import List
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from Bio import SeqIO
from phagecommander import Gene
import phagecommander.GuiWidgets
from phagecommander.Utilities import ThreadData, ProdigalRelease, Aragorn
from phagecommander.Utilities.Tools import *
import platform # (GRyde) Needed to disable Glimmer box for Windows

APP_NAME = 'Phage Commander'

# mappings of tool names to appropriate methods
# [queryMethod, parseMethod]
TOOL_METHODS = {GENEMARK: [Gene.GeneFile.genemark_query,
                           Gene.GeneParse.parse_genemark],
                HMM: [Gene.GeneFile.genemarkhmm_query,
                      Gene.GeneParse.parse_genemarkHmm],
                HEURISTIC: [Gene.GeneFile.genemark_heuristic_query,
                            Gene.GeneParse.parse_genemarkHeuristic],
                GENEMARKS: [Gene.GeneFile.genemarks_query,
                            Gene.GeneParse.parse_genemarkS],
                GENEMARKS2: [Gene.GeneFile.genemarks2_query,
                             Gene.GeneParse.parse_genemarkS2],
                GLIMMER: [Gene.GeneFile.glimmer_query,
                          Gene.GeneParse.parse_glimmer],
                PRODIGAL: [Gene.GeneFile.prodigal_query,
                           Gene.GeneParse.parse_prodigal],
                RAST: [Gene.GeneFile.rastQuery,
                       Gene.GeneParse.parse_rast],
                METAGENE: [Gene.GeneFile.metageneQuery,
                           Gene.GeneParse.parse_metagene],
                ARAGORN: [Gene.GeneFile.aragornQuery,
                          Gene.GeneParse.parse_aragorn]}


class QueryData:
    """
    Class for representing tool/species selections
    """

    def __init__(self):
        # tools to call
        self.tools = {key: True for key in TOOL_NAMES}
        # species of the DNA sequence
        self.species = ''
        # path of the DNA file
        self.fileName = ''
        # tool data
        # Key - tool (from TOOL_NAMES)
        # Value - List of Genes
        self.toolData = dict()
        # sequence
        self.sequence = ''
        # RAST related information
        self.rastUser = ''
        self.rastPass = ''
        self.rastJobID = None

    def wipeUserCredentials(self):
        """
        Deletes any data relating to a RAST query
        """
        self.rastJobID = None
        self.rastUser = None
        self.rastPass = None


class ColorTable(QWidget):
    CELL_COLOR_SETTING = 'TABLE/cell_color/'
    MAJORITY_TEXT_SETTING = 'TABLE/majority_text_color/'
    # (GRyde) ************************************************************************** start
    MAJORITY_TEXT_SECOND_SETTING = 'TABLE/majority_text_second_color/'
    MAJORITY_TEXT_THIRD_SETTING = 'TABLE/majority_text_third_color/'
    MAJORITY_TEXT_FOURTH_SETTING = 'TABLE/majority_text_fourth_color/'
    MAJORITY_TEXT_FIFTH_SETTING = 'TABLE/majority_text_fifth_color/'
    MAJORITY_TEXT_SIXTH_SETTING = 'TABLE/majority_text_sixth_color/'
    MAJORITY_TEXT_SEVENTH_SETTING = 'TABLE/majority_text_seventh_color/'
    MAJORITY_TEXT_EIGHTH_SETTING = 'TABLE/majority_text_eighth_color/'
    # (GRyde) ************************************************************************** end
    MINORITY_TEXT_SETTING = 'TABLE/minority_text_color/' # Need to change minority references to Ninth color
    _TABLE_COLUMN_HEADERS = [
        'Cell Color',
        'Majority Text',
        '2nd Text Color',
        '3rd Text Color',
        '4th Text Color',
        '5th Text Color',
        '6th Text Color',
        '7th Text Color',
        '8th Text Color',
        '9th Text Color'
    ]
    # Original code _TABLE_COLUMN_HEADERS = ['Cell Color', 'Majority Text', 'Minority Text']
    _TABLE_MAJORITY_MINORITY_DEFAULT_TEXT = '4099'
    _CELL_COLOR_COLUMN = 0
    _MAJORITY_TEXT_COLUMN = 1
    # (GRyde) ************************************************************************** start
    _SECOND_COLOR_TEXT_COLUMN = 2
    _THIRD_COLOR_TEXT_COLUMN = 3
    _FOURTH_COLOR_TEXT_COLUMN = 4
    _FIFTH_COLOR_TEXT_COLUMN = 5
    _SIXTH_COLOR_TEXT_COLUMN = 6
    _SEVENTH_COLOR_TEXT_COLUMN = 7
    _EIGHTH_COLOR_TEXT_COLUMN = 8
    # (GRyde) ************************************************************************** end
    _MINORITY_TEXT_COLUMN = 9 # Originally listed as column 2 before adding additional colors
    _DEFAULT_CELL_COLORS = [
        (255, 255, 255),
        (218, 238, 243),
        (183, 222, 232),
        (146, 205, 220),
        (49, 134, 155),
        (33, 89, 103),
        (21, 59, 68),
        (2, 47, 58),
        (1, 37, 56)
    ]
    _DEFAULT_MAJORITY_COLORS = [
        (0, 0, 0),
        (0, 0, 0),
        (0, 0, 0),
        (0, 0, 0),
        (255, 255, 255),
        (255, 255, 255),
        (255, 255, 255),
        (255, 255, 255),
        (255, 255, 255)
    ]
    # (GRyde) ************************************************************************** start
    # Since the nth color cannot appear unless called by at least n programs, all default colors set to black/white until
    # row n for nth Text Color
    _DEFAULT_MAJORITY_SECOND_COLORS = [
        (0, 0, 0),
        (255, 85, 0),
        (255, 85, 0),
        (255, 85, 0),
        (255, 170, 0),
        (255, 170, 0),
        (255, 170, 0),
        (255, 170, 0),
        (255, 170, 0)
    ]
    _DEFAULT_MAJORITY_THIRD_COLORS = [
        (0, 0, 0),
        (0, 0, 0),
        (0, 120, 0),
        (0, 120, 0),
        (170, 255, 170),
        (170, 255, 170),
        (170, 255, 170),
        (170, 255, 170),
        (170, 255, 170)
    ]
    _DEFAULT_MAJORITY_FOURTH_COLORS = [
        (0, 0, 0),
        (0, 0, 0),
        (0, 0, 0),
        (255, 0, 255),
        (255, 170, 255),
        (255, 170, 255),
        (255, 170, 255),
        (255, 170, 255),
        (255, 170, 255)
    ]
    _DEFAULT_MAJORITY_FIFTH_COLORS = [
        (0, 0, 0),
        (0, 0, 0),
        (0, 0, 0),
        (0, 0, 0),
        (255, 255, 0),
        (255, 255, 0),
        (255, 255, 0),
        (255, 255, 0),
        (255, 255, 0)
    ]
    _DEFAULT_MAJORITY_SIXTH_COLORS = [
        (0, 0, 0),
        (0, 0, 0),
        (0, 0, 0),
        (0, 0, 0),
        (255, 255, 255),
        (0, 255, 255),
        (0, 255, 255),
        (0, 255, 255),
        (0, 255, 255)
    ]
    _DEFAULT_MAJORITY_SEVENTH_COLORS = [
        (0, 0, 0),
        (0, 0, 0),
        (0, 0, 0),
        (0, 0, 0),
        (255, 255, 255),
        (255, 255, 255),
        (170, 170, 255),
        (170, 170, 255),
        (170, 170, 255)
    ]
    _DEFAULT_MAJORITY_EIGHTH_COLORS = [
        (0, 0, 0),
        (0, 0, 0),
        (0, 0, 0),
        (0, 0, 0),
        (255, 255, 255),
        (255, 255, 255),
        (255, 255, 255),
        (255, 85, 127),
        (255, 85, 127)
    ]
    # (GRyde) ************************************************************************** end
    _DEFAULT_MINORITY_COLORS = [
        (0, 0, 0),
        (0, 0, 0),
        (0, 0, 0),
        (0, 0, 0),
        (255, 255, 255),
        (255, 255, 255),
        (255, 255, 255),
        (255, 255, 255),
        (170, 170, 0)
    ]

    def __init__(self, settings, parent=None):
        super(ColorTable, self).__init__(parent)
        self.settings = settings

        layout = QVBoxLayout()

        # WIDGETS ------------------------------------------------------------------------------------------------------
        # Color Selection Table
        self.tableWidget = QTableWidget()
        self.tableWidget.setRowCount(len(GENE_TOOLS))
        self.tableWidget.setColumnCount(len(self._TABLE_COLUMN_HEADERS))
        self.tableWidget.setHorizontalHeaderLabels(self._TABLE_COLUMN_HEADERS)
        self.tableWidget.horizontalHeader()
        self.tableWidget.setSelectionMode(QTableWidget.NoSelection)
        self.tableWidget.cellClicked.connect(self.tableClick)
        self.tableWidget.setCornerButtonEnabled(False)
        self.tableWidget.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.tableWidget.horizontalScrollBar().setDisabled(True)
        self.tableWidget.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.tableWidget.verticalScrollBar().setDisabled(True)
        columnWidth = self.tableWidget.columnWidth(0)
        vertHeaderWidth = self.tableWidget.verticalHeader().width()
        self.tableWidget.setMinimumWidth(vertHeaderWidth + (10 * columnWidth)) # Originally (3 * columnWidth))
        self.tableWidget.setMaximumWidth(vertHeaderWidth + (10 * columnWidth)) # Originally (3 * columnWidth))

        # insert items into table
        tableHeight = self.tableWidget.horizontalHeader().height()
        for i in range(len(GENE_TOOLS)):
            self.tableWidget.setRowHeight(i, 20)
            tableHeight += self.tableWidget.rowHeight(i)
            item = QTableWidgetItem()
            majorityItem = QTableWidgetItem(self._TABLE_MAJORITY_MINORITY_DEFAULT_TEXT)
            # (GRyde) ****************************************************************** start
            secondTextItem = QTableWidgetItem(self._TABLE_MAJORITY_MINORITY_DEFAULT_TEXT)
            thirdTextItem = QTableWidgetItem(self._TABLE_MAJORITY_MINORITY_DEFAULT_TEXT)
            fourthTextItem = QTableWidgetItem(self._TABLE_MAJORITY_MINORITY_DEFAULT_TEXT)
            fifthTextItem = QTableWidgetItem(self._TABLE_MAJORITY_MINORITY_DEFAULT_TEXT)
            sixthTextItem = QTableWidgetItem(self._TABLE_MAJORITY_MINORITY_DEFAULT_TEXT)
            seventhTextItem = QTableWidgetItem(self._TABLE_MAJORITY_MINORITY_DEFAULT_TEXT)
            eighthTextItem = QTableWidgetItem(self._TABLE_MAJORITY_MINORITY_DEFAULT_TEXT)
            # (GRyde) ****************************************************************** end
            minorityItem = QTableWidgetItem(self._TABLE_MAJORITY_MINORITY_DEFAULT_TEXT)

            # set default cell color if no setting is found
            colorSetting = self.settings.value(self.CELL_COLOR_SETTING + str(i))
            if colorSetting is not None:
                colors = [int(num) for num in colorSetting.split(' ')]
                color = QColor(*colors)
            else:
                color = QColor(*self._DEFAULT_CELL_COLORS[i])
                colorStr = [str(num) for num in self._DEFAULT_CELL_COLORS[i]]
                self.settings.setValue(self.CELL_COLOR_SETTING + str(i), ' '.join(colorStr))
            item.setBackground(color)
            majorityItem.setBackground(color)
            # (GRyde) ****************************************************************** start
            secondTextItem.setBackground(color)
            thirdTextItem.setBackground(color)
            fourthTextItem.setBackground(color)
            fifthTextItem.setBackground(color)
            sixthTextItem.setBackground(color)
            seventhTextItem.setBackground(color)
            eighthTextItem.setBackground(color)
            # (GRyde) ****************************************************************** end
            minorityItem.setBackground(color)

            # set default majority color if no setting is found
            majoritySetting = self.settings.value(self.MAJORITY_TEXT_SETTING + str(i))
            if majoritySetting is not None:
                colors = [int(num) for num in majoritySetting.split(' ')]
                color = QColor(*colors)
                majorityItem.setForeground(color)
            else:
                defaultColor = QColor(*self._DEFAULT_MAJORITY_COLORS[i])
                majorityItem.setForeground(defaultColor)
                defaultColorStr = ' '.join([str(x) for x in defaultColor.getRgb()[:3]])
                self.settings.setValue(self.MAJORITY_TEXT_SETTING + str(i), defaultColorStr)
                
            # (GRyde) ****************************************************************** start
            # set default second text color if no setting is found
            secondTextColorSetting = self.settings.value(self.MAJORITY_TEXT_SECOND_SETTING + str(i))
            if secondTextColorSetting is not None:
                colors = [int(num) for num in secondTextColorSetting.split(' ')]
                color = QColor(*colors)
                secondTextItem.setForeground(color)
            else:
                defaultColor = QColor(*self._DEFAULT_MAJORITY_SECOND_COLORS[i])
                secondTextItem.setForeground(defaultColor)
                defaultColorStr = ' '.join([str(x) for x in defaultColor.getRgb()[:3]])
                self.settings.setValue(self.MAJORITY_TEXT_SECOND_SETTING + str(i), defaultColorStr)
                
            # set default third text color if no setting is found
            thirdTextColorSetting = self.settings.value(self.MAJORITY_TEXT_THIRD_SETTING + str(i))
            if thirdTextColorSetting is not None:
                colors = [int(num) for num in thirdTextColorSetting.split(' ')]
                color = QColor(*colors)
                thirdTextItem.setForeground(color)
            else:
                defaultColor = QColor(*self._DEFAULT_MAJORITY_THIRD_COLORS[i])
                thirdTextItem.setForeground(defaultColor)
                defaultColorStr = ' '.join([str(x) for x in defaultColor.getRgb()[:3]])
                self.settings.setValue(self.MAJORITY_TEXT_THIRD_SETTING + str(i), defaultColorStr)
                
            # set default fourth text color if no setting is found
            fourthTextColorSetting = self.settings.value(self.MAJORITY_TEXT_FOURTH_SETTING + str(i))
            if fourthTextColorSetting is not None:
                colors = [int(num) for num in fourthTextColorSetting.split(' ')]
                color = QColor(*colors)
                fourthTextItem.setForeground(color)
            else:
                defaultColor = QColor(*self._DEFAULT_MAJORITY_FOURTH_COLORS[i])
                fourthTextItem.setForeground(defaultColor)
                defaultColorStr = ' '.join([str(x) for x in defaultColor.getRgb()[:3]])
                self.settings.setValue(self.MAJORITY_TEXT_FOURTH_SETTING + str(i), defaultColorStr)
                
            # set default fifth text color if no setting is found
            fifthTextColorSetting = self.settings.value(self.MAJORITY_TEXT_FIFTH_SETTING + str(i))
            if fifthTextColorSetting is not None:
                colors = [int(num) for num in fifthTextColorSetting.split(' ')]
                color = QColor(*colors)
                fifthTextItem.setForeground(color)
            else:
                defaultColor = QColor(*self._DEFAULT_MAJORITY_FIFTH_COLORS[i])
                fifthTextItem.setForeground(defaultColor)
                defaultColorStr = ' '.join([str(x) for x in defaultColor.getRgb()[:3]])
                self.settings.setValue(self.MAJORITY_TEXT_FIFTH_SETTING + str(i), defaultColorStr)
                
            # set default sixth text color if no setting is found
            sixthTextColorSetting = self.settings.value(self.MAJORITY_TEXT_SIXTH_SETTING + str(i))
            if sixthTextColorSetting is not None:
                colors = [int(num) for num in sixthTextColorSetting.split(' ')]
                color = QColor(*colors)
                sixthTextItem.setForeground(color)
            else:
                defaultColor = QColor(*self._DEFAULT_MAJORITY_SIXTH_COLORS[i])
                sixthTextItem.setForeground(defaultColor)
                defaultColorStr = ' '.join([str(x) for x in defaultColor.getRgb()[:3]])
                self.settings.setValue(self.MAJORITY_TEXT_SIXTH_SETTING + str(i), defaultColorStr)
                
            # set default seventh text color if no setting is found
            seventhTextColorSetting = self.settings.value(self.MAJORITY_TEXT_SEVENTH_SETTING + str(i))
            if seventhTextColorSetting is not None:
                colors = [int(num) for num in seventhTextColorSetting.split(' ')]
                color = QColor(*colors)
                seventhTextItem.setForeground(color)
            else:
                defaultColor = QColor(*self._DEFAULT_MAJORITY_SEVENTH_COLORS[i])
                seventhTextItem.setForeground(defaultColor)
                defaultColorStr = ' '.join([str(x) for x in defaultColor.getRgb()[:3]])
                self.settings.setValue(self.MAJORITY_TEXT_SEVENTH_SETTING + str(i), defaultColorStr)
                
            # set default eighth text color if no setting is found
            eighthTextColorSetting = self.settings.value(self.MAJORITY_TEXT_EIGHTH_SETTING + str(i))
            if eighthTextColorSetting is not None:
                colors = [int(num) for num in eighthTextColorSetting.split(' ')]
                color = QColor(*colors)
                eighthTextItem.setForeground(color)
            else:
                defaultColor = QColor(*self._DEFAULT_MAJORITY_EIGHTH_COLORS[i])
                eighthTextItem.setForeground(defaultColor)
                defaultColorStr = ' '.join([str(x) for x in defaultColor.getRgb()[:3]])
                self.settings.setValue(self.MAJORITY_TEXT_EIGHTH_SETTING + str(i), defaultColorStr)
                
            # (GRyde) ****************************************************************** end
                
            # set default minority color if no setting is found
            minoritySetting = self.settings.value(self.MINORITY_TEXT_SETTING + str(i))
            if minoritySetting is not None:
                colors = [int(num) for num in minoritySetting.split(' ')]
                color = QColor(*colors)
                minorityItem.setForeground(color)
            else:
                defaultColor = QColor(*self._DEFAULT_MINORITY_COLORS[i])
                minorityItem.setForeground(defaultColor)
                defaultColorStr = ' '.join([str(x) for x in defaultColor.getRgb()[:3]])
                self.settings.setValue(self.MINORITY_TEXT_SETTING + str(i), defaultColorStr)

            # add items to table
            self.tableWidget.setItem(i, self._CELL_COLOR_COLUMN, item)
            self.tableWidget.setItem(i, self._MAJORITY_TEXT_COLUMN, majorityItem)
            # (GRyde) ****************************************************************** start
            self.tableWidget.setItem(i, self._SECOND_COLOR_TEXT_COLUMN, secondTextItem)
            self.tableWidget.setItem(i, self._THIRD_COLOR_TEXT_COLUMN, thirdTextItem)
            self.tableWidget.setItem(i, self._FOURTH_COLOR_TEXT_COLUMN, fourthTextItem)
            self.tableWidget.setItem(i, self._FIFTH_COLOR_TEXT_COLUMN, fifthTextItem)
            self.tableWidget.setItem(i, self._SIXTH_COLOR_TEXT_COLUMN, sixthTextItem)
            self.tableWidget.setItem(i, self._SEVENTH_COLOR_TEXT_COLUMN, seventhTextItem)
            self.tableWidget.setItem(i, self._EIGHTH_COLOR_TEXT_COLUMN, eighthTextItem)
            # (GRyde) ****************************************************************** end
            self.tableWidget.setItem(i, self._MINORITY_TEXT_COLUMN, minorityItem)
            majorityItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            # (GRyde) ****************************************************************** start
            secondTextItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            thirdTextItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            fourthTextItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            fifthTextItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            sixthTextItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            seventhTextItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            eighthTextItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            # (GRyde) ****************************************************************** end
            minorityItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
        self.tableWidget.setMaximumHeight(tableHeight)
        self.tableWidget.setMinimumHeight(tableHeight)

        # reset button
        resetButton = QPushButton('Reset ALL')
        resetButton.clicked.connect(self.resetToDefaultAll)
        layout.addWidget(self.tableWidget)
        layout.addWidget(resetButton)

        self.setLayout(layout)

    @pyqtSlot(int, int)
    def tableClick(self, row, column):
        """
        Called when an item in the table is clicked
        Performs the appropriate action on the item based on the column of item
        :param row:  int
        :param column: int
        """
        if column == self._CELL_COLOR_COLUMN:
            self.changeCellColor(row, column)
        if column == self._MAJORITY_TEXT_COLUMN or column == self._MINORITY_TEXT_COLUMN:
            self.changeTextColor(row, column)
        # (GRyde) ********************************************************************** start
        if (column == self._SECOND_COLOR_TEXT_COLUMN or column == self._THIRD_COLOR_TEXT_COLUMN or column == self._FOURTH_COLOR_TEXT_COLUMN or
            column == self._FIFTH_COLOR_TEXT_COLUMN or column == self._SIXTH_COLOR_TEXT_COLUMN or column == self._SEVENTH_COLOR_TEXT_COLUMN or
            column == self._EIGHTH_COLOR_TEXT_COLUMN):
            self.changeTextColor(row, column)
        # (GRyde) ********************************************************************** end

    def changeTextColor(self, row, column):
        """
        Changes the text color of the cell and saves the color to settings
        :param row: int
        :param column: int
        """
        tableItem = self.tableWidget.item(row, column)
        colorDialog = QColorDialog()
        currColor = tableItem.foreground().color()
        color = colorDialog.getColor(currColor, self, 'Select Text Color')
        if color.isValid():
            tableItem.setForeground(color)
            colorStr = ' '.join([str(x) for x in color.getRgb()[:3]])
            if column == self._MAJORITY_TEXT_COLUMN:
                self.settings.setValue(self.MAJORITY_TEXT_SETTING + str(row), colorStr)
            # (GRyde) ****************************************************************** start
            if column == self._SECOND_COLOR_TEXT_COLUMN:
                self.settings.setValue(self.MAJORITY_TEXT_SECOND_SETTING + str(row), colorStr)
            if column == self._THIRD_COLOR_TEXT_COLUMN:
                self.settings.setValue(self.MAJORITY_TEXT_THIRD_SETTING + str(row), colorStr)
            if column == self._FOURTH_COLOR_TEXT_COLUMN:
                self.settings.setValue(self.MAJORITY_TEXT_FOURTH_SETTING + str(row), colorStr)
            if column == self._FIFTH_COLOR_TEXT_COLUMN:
                self.settings.setValue(self.MAJORITY_TEXT_FIFTH_SETTING + str(row), colorStr)
            if column == self._SIXTH_COLOR_TEXT_COLUMN:
                self.settings.setValue(self.MAJORITY_TEXT_SIXTH_SETTING + str(row), colorStr)
            if column == self._SEVENTH_COLOR_TEXT_COLUMN:
                self.settings.setValue(self.MAJORITY_TEXT_SEVENTH_SETTING + str(row), colorStr)
            if column == self._EIGHTH_COLOR_TEXT_COLUMN:
                self.settings.setValue(self.MAJORITY_TEXT_EIGHTH_SETTING + str(row), colorStr)
            # (GRyde) ****************************************************************** end
            if column == self._MINORITY_TEXT_COLUMN:
                self.settings.setValue(self.MINORITY_TEXT_SETTING + str(row), colorStr)

    def changeCellColor(self, row, column):
        """
        Changes the color of the row and saves color to setting
        :param row: int
        :param column: int
        """
        item = self.tableWidget.item(row, column)
        majorityItem = self.tableWidget.item(row, column + 1)
        # (GRyde) ********************************************************************** start
        secondTextItem = self.tableWidget.item(row, column + 2)
        thirdTextItem = self.tableWidget.item(row, column + 3)
        fourthTextItem = self.tableWidget.item(row, column + 4)
        fifthTextItem = self.tableWidget.item(row, column + 5)
        sixthTextItem = self.tableWidget.item(row, column + 6)
        seventhTextItem = self.tableWidget.item(row, column + 7)
        eighthTextItem = self.tableWidget.item(row, column + 8)
        # (GRyde) ********************************************************************** end
        minorityItem = self.tableWidget.item(row, column + 9) # Originally column + 2
        colorDialog = QColorDialog()
        currColor = item.background().color()
        color = colorDialog.getColor(currColor, self, 'Select Cell Color')
        if color.isValid():
            item.setBackground(color)
            majorityItem.setBackground(color)
            # (GRyde) ****************************************************************** start
            secondTextItem.setBackground(color)
            thirdTextItem.setBackground(color)
            fourthTextItem.setBackground(color)
            fifthTextItem.setBackground(color)
            sixthTextItem.setBackground(color)
            seventhTextItem.setBackground(color)
            eighthTextItem.setBackground(color)
            # (GRyde) ****************************************************************** end
            minorityItem.setBackground(color)
            colorStr = ' '.join([str(x) for x in color.getRgb()[:3]])
            self.settings.setValue(self.CELL_COLOR_SETTING + str(row), colorStr)

    def resetToDefaultAll(self):
        """
        Resets all rows in ColorSettings dialog to default colors
        """
        # prompt user if they wish to reset
        response = QMessageBox.warning(self,
                                       'Reset to Default',
                                       'Reset to Default? - Existing settings will be reset',
                                       QMessageBox.Ok | QMessageBox.Cancel)

        if response == QMessageBox.Ok:
            for row in range(self.tableWidget.rowCount()):
                # set majority text color
                majorityItem = self.tableWidget.item(row, self._MAJORITY_TEXT_COLUMN)
                majorityColor = QColor(*self._DEFAULT_MAJORITY_COLORS[row])
                majorityItem.setForeground(majorityColor)
                majorityColorStr = ' '.join([str(x) for x in majorityColor.getRgb()[:3]])
                self.settings.setValue(self.MAJORITY_TEXT_SETTING + str(row), majorityColorStr)
                
                # (GRyde) ************************************************************** start
                # Works for resetting colors for now though!
                
                
                # MAJORITY SECOND COLOR
                secondTextItem = self.tableWidget.item(row, self._SECOND_COLOR_TEXT_COLUMN)
                secondTextColor = QColor(*self._DEFAULT_MAJORITY_SECOND_COLORS[row])
                secondTextItem.setForeground(secondTextColor)
                majoritySecondColorStr = ' '.join([str(x) for x in secondTextColor.getRgb()[:3]])
                self.settings.setValue(self.MAJORITY_TEXT_SECOND_SETTING + str(row), majoritySecondColorStr)
                
                # MAJORITY THIRD COLOR
                thirdTextItem = self.tableWidget.item(row, self._THIRD_COLOR_TEXT_COLUMN)
                thirdTextColor = QColor(*self._DEFAULT_MAJORITY_THIRD_COLORS[row])
                thirdTextItem.setForeground(thirdTextColor)
                majorityThirdColorStr = ' '.join([str(x) for x in thirdTextColor.getRgb()[:3]])
                self.settings.setValue(self.MAJORITY_TEXT_THIRD_SETTING + str(row), majorityThirdColorStr)
                
                # MAJORITY FOURTH COLOR
                fourthTextItem = self.tableWidget.item(row, self._FOURTH_COLOR_TEXT_COLUMN)
                fourthTextColor = QColor(*self._DEFAULT_MAJORITY_FOURTH_COLORS[row])
                fourthTextItem.setForeground(fourthTextColor)
                majorityFourthColorStr = ' '.join([str(x) for x in fourthTextColor.getRgb()[:3]])
                self.settings.setValue(self.MAJORITY_TEXT_FOURTH_SETTING + str(row), majorityFourthColorStr)
                
                # MAJORITY FIFTH COLOR
                fifthTextItem = self.tableWidget.item(row, self._FIFTH_COLOR_TEXT_COLUMN)
                fifthTextColor = QColor(*self._DEFAULT_MAJORITY_FIFTH_COLORS[row])
                fifthTextItem.setForeground(fifthTextColor)
                majorityFifthColorStr = ' '.join([str(x) for x in fifthTextColor.getRgb()[:3]])
                self.settings.setValue(self.MAJORITY_TEXT_FIFTH_SETTING + str(row), majorityFifthColorStr)
                
                # MAJORITY SIXTH COLOR
                sixthTextItem = self.tableWidget.item(row, self._SIXTH_COLOR_TEXT_COLUMN)
                sixthTextColor = QColor(*self._DEFAULT_MAJORITY_SIXTH_COLORS[row])
                sixthTextItem.setForeground(sixthTextColor)
                majoritySixthColorStr = ' '.join([str(x) for x in sixthTextColor.getRgb()[:3]])
                self.settings.setValue(self.MAJORITY_TEXT_SIXTH_SETTING + str(row), majoritySixthColorStr)
                
                # MAJORITY SEVENTH COLOR
                seventhTextItem = self.tableWidget.item(row, self._SEVENTH_COLOR_TEXT_COLUMN)
                seventhTextColor = QColor(*self._DEFAULT_MAJORITY_SEVENTH_COLORS[row])
                seventhTextItem.setForeground(seventhTextColor)
                majoritySeventhColorStr = ' '.join([str(x) for x in seventhTextColor.getRgb()[:3]])
                self.settings.setValue(self.MAJORITY_TEXT_SEVENTH_SETTING + str(row), majoritySeventhColorStr)
                
                # MAJORITY EIGHTH COLOR
                eighthTextItem = self.tableWidget.item(row, self._EIGHTH_COLOR_TEXT_COLUMN)
                eighthTextColor = QColor(*self._DEFAULT_MAJORITY_EIGHTH_COLORS[row])
                eighthTextItem.setForeground(eighthTextColor)
                majorityEighthColorStr = ' '.join([str(x) for x in eighthTextColor.getRgb()[:3]])
                self.settings.setValue(self.MAJORITY_TEXT_EIGHTH_SETTING + str(row), majorityEighthColorStr)
                
                # (GRyde) ************************************************************** end

                # set minority text color
                minorityItem = self.tableWidget.item(row, self._MINORITY_TEXT_COLUMN)
                minorityColor = QColor(*self._DEFAULT_MINORITY_COLORS[row])
                minorityItem.setForeground(minorityColor)
                minorityColorStr = ' '.join([str(x) for x in minorityColor.getRgb()[:3]])
                self.settings.setValue(self.MINORITY_TEXT_SETTING + str(row), minorityColorStr)

                # set cell color
                cellColorItem = self.tableWidget.item(row, self._CELL_COLOR_COLUMN)
                cellColor = QColor(*self._DEFAULT_CELL_COLORS[row])
                cellColorItem.setBackground(cellColor)
                majorityItem.setBackground(cellColor)
                # (GRyde) ************************************************************** start
                secondTextItem.setBackground(cellColor)
                thirdTextItem.setBackground(cellColor)
                fourthTextItem.setBackground(cellColor)
                fifthTextItem.setBackground(cellColor)
                sixthTextItem.setBackground(cellColor)
                seventhTextItem.setBackground(cellColor)
                eighthTextItem.setBackground(cellColor)
                # (GRyde) ************************************************************** end
                minorityItem.setBackground(cellColor)
                cellColorStr = ' '.join([str(x) for x in cellColor.getRgb()[:3]])
                self.settings.setValue(self.CELL_COLOR_SETTING + str(row), cellColorStr)

    @staticmethod
    def checkDefaultSettings(settings):
        """
        Checks if the default color settings exist in the QSettings. If not, populates them
        :param settings: QSettings
        """
        cellColorSettings = [settings.value(ColorTable.CELL_COLOR_SETTING + str(i)) for i in range(len(GENE_TOOLS))]
        majorityColorSettings = [settings.value(ColorTable.MAJORITY_TEXT_SETTING + str(i)) for i in
                                 range(len(GENE_TOOLS))]
        # (GRyde) ********************************************************************** start
        majoritySecondColorSettings = [settings.value(ColorTable.MAJORITY_TEXT_SECOND_SETTING + str(i)) for i in
                                       range(len(GENE_TOOLS))]
        majorityThirdColorSettings = [settings.value(ColorTable.MAJORITY_TEXT_THIRD_SETTING + str(i)) for i in
                                      range(len(GENE_TOOLS))]
        majorityFourthColorSettings = [settings.value(ColorTable.MAJORITY_TEXT_FOURTH_SETTING + str(i)) for i in
                                       range(len(GENE_TOOLS))]
        majorityFifthColorSettings = [settings.value(ColorTable.MAJORITY_TEXT_FIFTH_SETTING + str(i)) for i in
                                      range(len(GENE_TOOLS))]
        majoritySixthColorSettings = [settings.value(ColorTable.MAJORITY_TEXT_SIXTH_SETTING + str(i)) for i in
                                      range(len(GENE_TOOLS))]
        majoritySeventhColorSettings = [settings.value(ColorTable.MAJORITY_TEXT_SEVENTH_SETTING + str(i)) for i in
                                        range(len(GENE_TOOLS))]
        majorityEighthColorSettings = [settings.value(ColorTable.MAJORITY_TEXT_EIGHTH_SETTING + str(i)) for i in
                                       range(len(GENE_TOOLS))]
        # (GRyde) ********************************************************************** end
        minorityColorSettings = [settings.value(ColorTable.MINORITY_TEXT_SETTING + str(i)) for i in
                                 range(len(GENE_TOOLS))]

        if None in cellColorSettings or None in majorityColorSettings or None in minorityColorSettings:
            ColorTable._setDefaultSettings(settings)
        # (GRyde) ********************************************************************** start
        # Definitely could clean this part up, now that I know this is what gets fix to work I should make these part of the
        # original 'if' statement above
        # Maybe add all the settings items to a list and iterate through it?
        if (None in majoritySecondColorSettings or None in majorityThirdColorSettings or None in majorityFourthColorSettings or
            None in majorityFifthColorSettings or None in majoritySixthColorSettings or None in majoritySeventhColorSettings or
            None in majorityEighthColorSettings):
            ColorTable._setDefaultSettings(settings)
        # (GRyde) ********************************************************************** end

    @staticmethod
    def _setDefaultSettings(settings):
        """
        Sets the settings related to cell colors to default values
        :param settings: QSettings object
        """
        for i in range(len(GENE_TOOLS)):
            # CELL COLORS
            defaultColorStr = ' '.join(str(color) for color in ColorTable._DEFAULT_CELL_COLORS[i])
            settings.setValue(ColorTable.CELL_COLOR_SETTING + str(i), defaultColorStr)

            # MAJORITY COLOR
            defaultColorStr = ' '.join(str(color) for color in ColorTable._DEFAULT_MAJORITY_COLORS[i])
            settings.setValue(ColorTable.MAJORITY_TEXT_SETTING + str(i), defaultColorStr)
            
            # (GRyde) ******************************************************************* start
            
            # MAJORITY SECOND COLOR
            defaultColorStr = ' '.join(str(color) for color in ColorTable._DEFAULT_MAJORITY_SECOND_COLORS[i])
            settings.setValue(ColorTable.MAJORITY_TEXT_SECOND_SETTING + str(i), defaultColorStr)
            
            # MAJORITY THIRD COLOR
            defaultColorStr = ' '.join(str(color) for color in ColorTable._DEFAULT_MAJORITY_THIRD_COLORS[i])
            settings.setValue(ColorTable.MAJORITY_TEXT_THIRD_SETTING + str(i), defaultColorStr)
            
            # MAJORITY FOURTH COLOR
            defaultColorStr = ' '.join(str(color) for color in ColorTable._DEFAULT_MAJORITY_FOURTH_COLORS[i])
            settings.setValue(ColorTable.MAJORITY_TEXT_FOURTH_SETTING + str(i), defaultColorStr)
            
            # MAJORITY FIFTH COLOR
            defaultColorStr = ' '.join(str(color) for color in ColorTable._DEFAULT_MAJORITY_FIFTH_COLORS[i])
            settings.setValue(ColorTable.MAJORITY_TEXT_FIFTH_SETTING + str(i), defaultColorStr)
            
            # MAJORITY SIXTH COLOR
            defaultColorStr = ' '.join(str(color) for color in ColorTable._DEFAULT_MAJORITY_SIXTH_COLORS[i])
            settings.setValue(ColorTable.MAJORITY_TEXT_SIXTH_SETTING + str(i), defaultColorStr)
            
            # MAJORITY SEVENTH COLOR
            defaultColorStr = ' '.join(str(color) for color in ColorTable._DEFAULT_MAJORITY_SEVENTH_COLORS[i])
            settings.setValue(ColorTable.MAJORITY_TEXT_SEVENTH_SETTING + str(i), defaultColorStr)
            
            # MAJORITY EIGHTH COLOR
            defaultColorStr = ' '.join(str(color) for color in ColorTable._DEFAULT_MAJORITY_EIGHTH_COLORS[i])
            settings.setValue(ColorTable.MAJORITY_TEXT_EIGHTH_SETTING + str(i), defaultColorStr)
            
            # (GRyde) ******************************************************************* end

            # MINORITY COLOR
            defaultColorStr = ' '.join(str(color) for color in ColorTable._DEFAULT_MINORITY_COLORS[i])
            settings.setValue(ColorTable.MINORITY_TEXT_SETTING + str(i), defaultColorStr)


class SettingsDialog(QDialog):
    """
    Dialog for Settings
    """

    def __init__(self, parent=None):
        super(SettingsDialog, self).__init__(parent)
        self.settings = QSettings(QSettings.IniFormat, QSettings.UserScope, APP_NAME, APP_NAME)

        # WIDGETS --------------------------------------------------------------------
        layout = QVBoxLayout()
        self.tabWidget = QTabWidget()
        self.tableTab = QWidget()

        # UI Initializations
        self.initTableTab()

        # Layout
        layout.addWidget(self.tabWidget)
        self.setLayout(layout)

        # Window Settings
        self.setWindowTitle('Settings')
        self.setWindowFlags(Qt.WindowCloseButtonHint)

    def initTableTab(self):
        self.tableTab = ColorTable(self.settings)
        self.tabWidget.addTab(self.tableTab, 'Table')


class NewFileDialog(QDialog):
    """
    Dialog shown when making a new query
    :returns: list of tools user selected to make queries to
    """

    _LAST_FASTA_FILE_LOCATION_SETTING = 'NEW_FILE_DIALOG/last_fasta_location'
    _RAST_USERNAME_SETTING = 'NEW_FILE_DIALOG/rast_username'
    _RAST_PASSWORD_SETTING = 'NEW_FILE_DIALOG/rast_password'

    def __init__(self, queryData, settings, prodigalPath=None, parent=None):
        """
        Initialize Dialog
        :param queryData: ToolSpecies object
        :param settings: QSettings
        :param prodigalPath: path to Prodigal binary
            * str or None
        :param parent: parent widget

        """
        super(NewFileDialog, self).__init__(parent)
        self.queryData = queryData
        self.settings = settings

        mainLayout = QVBoxLayout()
        checkBoxLayout = QGridLayout()
        speciesLayout = QHBoxLayout()
        dnaFileLayout = QHBoxLayout()
        buttonLayout = QHBoxLayout()

        # label font
        labelFont = QFont()
        labelFont.setUnderline(True)
        labelFont.setPointSize(12)

        # WIDGETS ----------------------------------------------------------------------------------
        # genemark boxes
        """
            (GRyde) Update to checkbox labels
            HMM > Host-trained GeneMark
            Heuristic > GeneMark w/ Heuristics
            GMS > GeneMark S
            GMS2 > GeneMark S2
        """
        genemarkLabel = QLabel('GeneMark')
        genemarkLabel.setFont(labelFont)
        gmBox = QCheckBox('GeneMark')
        self.hmmBox = QCheckBox('Host-trained GeneMark')
        heuristicBox = QCheckBox('GeneMark w/ Heuristics')
        gmsBox = QCheckBox('GeneMark S')
        gms2Box = QCheckBox('GeneMark S2')

        # glimmer box
        glimmerLabel = QLabel('Glimmer')
        glimmerLabel.setFont(labelFont)
        glimmerBox = QCheckBox('Glimmer')

        # prodigal box
        prodigalLabel = QLabel('Prodigal')
        prodigalLabel.setFont(labelFont)
        prodigalBox = QCheckBox('Prodigal')

        # rast box
        rastLabel = QLabel('RAST')
        rastLabel.setFont(labelFont)
        rastBox = QCheckBox('RAST')

        # metagene box
        METAGENE_LABEL_TEXT = 'Metagene'
        metageneLabel = QLabel(METAGENE_LABEL_TEXT)
        metageneLabel.setFont(labelFont)
        metageneBox = QCheckBox(METAGENE_LABEL_TEXT)

        # aragorn box
        ARAGORN_LABEL_TEXT = 'Aragorn'
        aragornLabel = QLabel(ARAGORN_LABEL_TEXT)
        aragornLabel.setFont(labelFont)
        aragornBox = QCheckBox('Aragorn (for tRNA)')
        
        # checkbox font
        checkBoxFont = QFont()
        checkBoxFont.setPointSize(12)
        
        # dictionary mapping tools to checkboxes
        self.toolCheckBoxes = dict()
        self.toolCheckBoxes['gm'] = gmBox
        self.toolCheckBoxes['hmm'] = self.hmmBox
        self.toolCheckBoxes['heuristic'] = heuristicBox
        self.toolCheckBoxes['gms'] = gmsBox
        self.toolCheckBoxes['gms2'] = gms2Box
        self.toolCheckBoxes['glimmer'] = glimmerBox
        self.toolCheckBoxes['prodigal'] = prodigalBox
        self.toolCheckBoxes['rast'] = rastBox
        self.toolCheckBoxes[METAGENE] = metageneBox
        self.toolCheckBoxes[ARAGORN] = aragornBox
        for box in self.toolCheckBoxes.values():
            # set all boxes to default to being checked
            # box.setChecked(True)
            box.setCheckState(Qt.Checked)
            # check to disable the species combobox on every click of a box
            box.stateChanged.connect(self.disableSpeciesCheck)
            # Set font size
            box.setFont(checkBoxFont)


        # check if Prodigal binary exists, if not, disable Prodigal
        if prodigalPath is None:
            prodigalBox.setCheckable(False)
            prodigalBox.setEnabled(False)
            
        
        # species combo box
        speciesLabel = QLabel('Host species (for host-trained GeneMark):')
        speciesLabel.setFont(labelFont)
        self.speciesComboBox = QComboBox()
        self.speciesComboBox.addItems(Gene.SPECIES)
        self.speciesComboBox.setMaximumWidth(550) # Originally 250
        
        # (GRyde) Testing disable of Glimmer box (works, just need to uncomment when ready to use)
        # platformCheck = platform.system()
        # platformCheck = platformCheck.lower()
        
        # if platformCheck == 'windows':
            # glimmerBox.setCheckState(Qt.Unchecked)
            # glimmerBox.setEnabled(False)
        

        # dna file input
        fileLabel = QLabel('Input genome (.fasta file):')
        fileLabel.setFont(labelFont)
        self.fileEdit = QLineEdit()
        fileButton = QPushButton('Select...')
        fileButton.clicked.connect(self.openFileDialog)

        # buttons
        self.queryButton = QPushButton('Run Phage Commander')
        self.queryButton.clicked.connect(self.accept)
        self.cancelButton = QPushButton('Cancel')
        self.cancelButton.clicked.connect(self.reject)

        # Widget Layout ----------------------------------------------------------------------------
        # (GRyde) It has been requested to remove the labels (underlined text). Will just comment then out here
        #         so re-adding is easy if necessary
        # genemark
        # (GRyde) Order of checkboxes has also been requested to be updated
        
        # rast
        #checkBoxLayout.addWidget(rastLabel, 3, 2)
        checkBoxLayout.addWidget(rastBox, 1, 0)
        
        # prodigal
        #checkBoxLayout.addWidget(prodigalLabel, 3, 1)
        checkBoxLayout.addWidget(prodigalBox, 1, 1)
        
        # metagene
        #checkBoxLayout.addWidget(metageneLabel, 5, 0)
        checkBoxLayout.addWidget(metageneBox, 1, 2)
       
        # glimmer
        #checkBoxLayout.addWidget(glimmerLabel, 3, 0)
        checkBoxLayout.addWidget(glimmerBox, 2, 0)
        
        #checkBoxLayout.addWidget(genemarkLabel, 0, 0)
        checkBoxLayout.addWidget(gmBox, 2, 1)
        checkBoxLayout.addWidget(heuristicBox, 2, 2)
        checkBoxLayout.addWidget(gmsBox, 3, 0)
        checkBoxLayout.addWidget(gms2Box, 3, 1)
        checkBoxLayout.addWidget(self.hmmBox, 3, 2)
        
        # aragorn
        #checkBoxLayout.addWidget(aragornLabel, 5, 1)
        #checkBoxLayout.addWidget(aragornBox, 4, 0)

        # species
        speciesLayout.addWidget(speciesLabel)
        # .addWidget(widget, stretch factor)
        speciesLayout.addWidget(self.speciesComboBox, 2)

        # file
        dnaFileLayout.addWidget(self.fileEdit)
        dnaFileLayout.addWidget(fileButton)

        # buttons
        buttonLayout.addWidget(self.queryButton)
        buttonLayout.addWidget(self.cancelButton)
        
        mainLayout.addLayout(checkBoxLayout)
        mainLayout.addLayout(speciesLayout)
        mainLayout.addWidget(aragornBox)
        mainLayout.addWidget(fileLabel)
        mainLayout.addLayout(dnaFileLayout)
        mainLayout.addLayout(buttonLayout)
        # (GRyde) ******************************************************* start
        # WINDOW SIZE
        self.setMinimumWidth(650)
        self.setMinimumHeight(250)
        # (GRyde) ******************************************************* end

        # Dialog Settings --------------------------------------------------------------------------
        self.setLayout(mainLayout)
        self.setWindowTitle('Select Gene Identification Programs')
        self.setWindowFlags(Qt.WindowCloseButtonHint | Qt.MSWindowsFixedSizeDialogHint)

    @pyqtSlot()
    def accept(self):
        """
        On Clicking "Query"
        """
        # reset toolData
        self.queryData.toolData = dict()

        # update tool calls based on user selection + allocates entry in toolData
        for key in self.queryData.tools.keys():
            self.queryData.tools[key] = self.toolCheckBoxes[key].isChecked()
            if self.queryData.tools[key] is True:
                self.queryData.toolData[key] = None

        # check if no tools were selected
        if True not in self.queryData.tools.values():
            QMessageBox.warning(self, 'Please select a tool',
                                'No tools are selected. Please choose at least one.')
            return

        # update species
        self.queryData.species = self.speciesComboBox.currentText()

        # check if dna file was given
        if self.fileEdit.text() == '':
            QMessageBox.warning(self, 'Missing DNA File',
                                'Please provide a fasta DNA file.')
            return

        # check if file exists
        if not os.path.isfile(self.fileEdit.text()):
            QMessageBox.warning(self, 'File Does not Exist',
                                'Selected DNA file does not exist.')
            return

        # if RAST was selected, prompt credential window
        if self.toolCheckBoxes['rast'].isChecked():
            credDialog = phagecommander.GuiWidgets.RastJobDialog(self.queryData)
            # if user exits window without submitting, do not query
            if not credDialog.exec_():
                return

        # update return values
        self.queryData.fileName = self.fileEdit.text()

        QDialog.accept(self)

    @pyqtSlot()
    def reject(self):
        """
        On Clicking "Cancel"
        """
        QDialog.reject(self)

    def openFileDialog(self):
        """
        Open a dialog for user to select DNA file
        """
        last_fasta_file_location = self.settings.value(self._LAST_FASTA_FILE_LOCATION_SETTING)
        # (GRyde) Request to filter only .fasta files
        file = QFileDialog.getOpenFileName(self, 'Open DNA File', last_fasta_file_location, "FASTA file (*.fasta)")

        # if file was chosen, set file line edit
        if file[0]:
            self.fileEdit.setText(file[0])
            # set new last_fasta_location
            fileFolder = os.path.split(file[0])[0]
            self.settings.setValue(self._LAST_FASTA_FILE_LOCATION_SETTING, fileFolder)

    def disableSpeciesCheck(self):
        """
        Disables the species comboBox if none of the selected tools require it
        * Only Hmm uses the species comboBox
        """
        if not self.hmmBox.isChecked():
            self.speciesComboBox.setDisabled(True)
        else:
            self.speciesComboBox.setDisabled(False)

    @staticmethod
    def checkDefaultSettings(settings):
        """
        Checks if required settings exist - if not, populates them
        :param settings: QSettings obj
        """
        # LAST FASTA FILE LOCATION
        if settings.value(NewFileDialog._LAST_FASTA_FILE_LOCATION_SETTING) is None:
            NewFileDialog._setDefaultSettings(settings)

    @staticmethod
    def _setDefaultSettings(settings):
        """
        Populates the default values for the required settings
        :param settings: QSettings obj
        """
        # LAST FASTA FILE LOCATION
        settings.setValue(NewFileDialog._LAST_FASTA_FILE_LOCATION_SETTING, '')


class QueryThread(QThread):
    """
    Thread for making performing the call to a gene prediction tool and parsing the data
    Returns the list of Genes via reference
    """

    def __init__(self, geneFile, tool, queryData, settings):
        """
        Constructor
        :param geneFile: GeneFile object of DNA file
        :param tool: tool to call
        :param settings: QSettings
            * See TOOL_NAMES global
        """
        super(QueryThread, self).__init__()

        self.tool = tool
        self.queryData = queryData
        self.geneFile = geneFile
        self.settings = settings

    def run(self):
        """
        Performs the query of the gene prediction tool and parses the output data
        :return: a list of Genes is returned through self.geneData
        """
        queryMethod = TOOL_METHODS[self.tool][0]
        parseMethod = TOOL_METHODS[self.tool][1]

        # perform query
        # if query is unsuccessful, return the error instead
        try:
            if self.tool == RAST:
                username = self.queryData.rastUser
                password = self.queryData.rastPass
                jobID = self.queryData.rastJobID
                queryMethod(self.geneFile, username, password, jobId=jobID)
            else:
                queryMethod(self.geneFile)
        except Exception as e:
            self.queryData.toolData[self.tool] = e
            return
            
        # (GRyde) Call to parse methods updated with third argument, which is length of gene sequence
        try:
            genes = parseMethod(self.geneFile.query_data[self.tool], identity=self.tool, totalLength=len(self.queryData.sequence))
        except Exception as e:
            self.queryData.toolData[self.tool] = e
            return
            

        """
            (GRyde) Original code
        # perform parsing of data
        try:
            genes = parseMethod(self.geneFile.query_data[self.tool], identity=self.tool)
        except Exception as e:
            self.queryData.toolData[self.tool] = e
            return
        """

        # update query object with genes
        self.queryData.toolData[self.tool] = genes

        # wipe RAST user creds
        if self.tool == RAST:
            self.queryData.wipeUserCredentials()


class QueryManager(QThread):
    """
    Thread for managing gene prediction tool calls
    """
    # signal emitted each time a querying thread returns
    progressSig = pyqtSignal()

    def __init__(self, queryData, settings):
        """
        Initializes and starts threads for each tool to be called
        :param queryData: QueryData object
        """
        super(QueryManager, self).__init__()

        # VARIABLES --------------------------------------------------------------------------------
        self.queryData = queryData
        self.settings = settings

        # create GeneFile
        self.geneFile = Gene.GeneFile(self.queryData.fileName, self.queryData.species,
                                      self.settings.value(GeneMain._PRODIGAL_BINARY_LOCATION_SETTING))

        # load sequence
        # with open(self.queryData.fileName) as seqFile:
        #     self.queryData.sequence = seqFile.read().split('\n')[1].lower()
        for seq_rec in SeqIO.parse(self.queryData.fileName, 'fasta'):
            self.queryData.sequence = seq_rec

        # THREAD ALLOCATIONS -----------------------------------------------------------------------
        self.threads = []
        for tool in self.queryData.tools:
            if self.queryData.tools[tool] is True:
                self.threads.append(QueryThread(self.geneFile, tool, self.queryData, self.settings))

        for thread in self.threads:
            thread.finished.connect(self.queryReturn)

        for thread in self.threads:
            thread.start()

    @pyqtSlot()
    def queryReturn(self):
        # emit progressSig to update progressBar
        self.progressSig.emit()

        # keep waiting until all calls have returned
        for tool in self.queryData.toolData:
            if self.queryData.toolData[tool] is None:
                return

        self.exit()

    def abort(self):
        self.exit()


class QueryDialog(QDialog):
    """
    Dialog for querying prediction tools
    """

    def __init__(self, queryData, settings, parent=None):
        super(QueryDialog, self).__init__(parent)

        self.queryData = queryData
        self.settings = settings

        mainLayout = QVBoxLayout()
        # WIDGETS ----------------------------------------------------------------------------------
        self.thread = QueryManager(queryData, self.settings)
        self.thread.finished.connect(self.queryStop)
        self.thread.progressSig.connect(self.updateProgress)

        self.progressBar = QProgressBar()
        # set progress max to the amount of tools to be queried
        self.progressBar.setMaximum(list(self.queryData.tools.values()).count(True))

        self.cancelButton = QPushButton('Cancel')
        self.cancelButton.clicked.connect(self.thread.abort)

        # WIDGET LAYOUT ----------------------------------------------------------------------------
        mainLayout.addWidget(self.progressBar)
        mainLayout.addWidget(self.cancelButton)
        self.setLayout(mainLayout)

        # SETTINGS ---------------------------------------------------------------------------------
        self.setWindowFlags(Qt.WindowTitleHint)
        self.setWindowTitle('Querying Tools...')

        self.progressBar.setValue(0)
        self.thread.start()

    @pyqtSlot()
    def updateProgress(self):
        """
        Increment the progress bar by one
        """
        self.progressBar.setValue(self.progressBar.value() + 1)

    @pyqtSlot()
    def queryStop(self):
        """
        Called when query thread stops - whether by finishing or user pressing cancel
        """
        # if successful query - display success message
        if self.progressBar.value() == self.progressBar.maximum():
            # check for any errors returned
            errors = dict()
            for tool in self.queryData.toolData:
                if isinstance(self.queryData.toolData[tool], Exception):
                    errors[tool] = self.queryData.toolData[tool]

            # errors exist
            if len(errors) != 0:
                # print out all tools and their errors
                errorStr = ['{}: {}'.format(tool.upper(), error) for tool, error in errors.items()]
                QMessageBox.information(self, 'Errors while Querying', '\n'.join(errorStr))
                QDialog.reject(self)
            else:
                # success
                QMessageBox.information(self, 'Done', 'Done! Query Successful')
                QDialog.accept(self)
        else:
            QDialog.reject(self)


class exportGenbankDialog(phagecommander.GuiWidgets.exportDialog):
    _LAST_GENBANK_LOCATION_SETTING = 'EXPORT_GENBANK_DIALOG/last_genbank_location'

    def __init__(self, queryData, settings, parent=None):
        super(exportGenbankDialog, self).__init__(queryData, settings, parent=parent)

        # WINDOW ---------------------------------------------------------------
        self.setWindowTitle('Export to Genbank')

    def saveFile(self):
        fileExtensions = ['Genbank (*.gb)',
                          'All Files (*.*)']
        saveFileLocation = self.settings.value(exportGenbankDialog._LAST_GENBANK_LOCATION_SETTING)
        saveFileName = QFileDialog.getSaveFileName(self,
                                                   'Save Genbank File As...',
                                                   saveFileLocation,
                                                   ';;'.join(fileExtensions))

        # if the user gave a file
        if saveFileName[0]:
            self.saveLineEdit.setText(saveFileName[0])

    def accept(self):
        """
        Called when user presses export
        """

        # put all Genes in one list
        filteredGenes = []
        for geneSet in self.queryData.toolData.values():
            filteredGenes.extend(geneSet)

        # filter based on user selection
        filteredGenes = Gene.GeneUtils.filterGenes(filteredGenes, self.getFilterFunction(), self.exportTRNA)
        # filtered genes is now List[List[Gene]]
        
        genesToExport = []
        # (GRyde) Adding methods for exporting based on longest gene or specific program
        
        if self.codonCurrentSelection == self._MOST_OCCURRENCES_TEXT:
            # find the most frequent Gene in each set of Genes
            for geneSet in filteredGenes:
                genesToExport.append(Gene.GeneUtils.findMostGeneOccurrences(geneSet))
                
        elif self.codonCurrentSelection == self._LONGEST_TEXT:
            # find the longest Gene in each set of Genes
            for geneSet in filteredGenes:
                genesToExport.append(Gene.GeneUtils.findLongestGene(geneSet))
                
        elif self.codonCurrentSelection == self._SPECIFIC_PROGRAM_TEXT:
            # find the gene called by a specific program (if not called by program, use most or longest in tie)
            programFilter = self.getSpecificProgram()
            for geneSet in filteredGenes:
                genesToExport.append(Gene.GeneUtils.useSpecificProgram(geneSet, programFilter))

        # output to file
        try:
            Gene.GeneUtils.genbankToFile(str(self.queryData.sequence.seq).lower(), genesToExport, self.saveFileName)
        except PermissionError as e:
            QMessageBox.warning(self,
                                'Permission Denied',
                                'Could not write to: \"{}\". Permission denied.'.format(self.saveFileName))
            # go back to dialogue
            return
        except Exception as e:
            QMessageBox.warning(self,
                                'Could Not Write to File',
                                'Could not write to: \"{}\".\n{}'.format(self.saveFileName, str(e)))
            # go back to dialogue
            return

        # save file location setting
        saveFileDir = os.path.split(self.saveFileName)[0]
        self.settings.setValue(exportGenbankDialog._LAST_GENBANK_LOCATION_SETTING, saveFileDir)

        QDialog.accept(self)

    @staticmethod
    def checkDefaultSettings(settings):
        """
        Checks if required settings exist - if not, populates them
        :param settings: QSettings obj
        """
        # Last save location
        if settings.value(exportGenbankDialog._LAST_GENBANK_LOCATION_SETTING) is None:
            exportGenbankDialog._setDefaultSettings(settings)

    @staticmethod
    def _setDefaultSettings(settings):
        """
        Populates the default values for the required settings
        :param settings: QSettings obj
        """
        # Last save location
        settings.setValue(exportGenbankDialog._LAST_GENBANK_LOCATION_SETTING, '')


class GeneMain(QMainWindow):
    """
    Main Window
    """

    _LAST_OPEN_FILE_LOCATION_SETTING = 'GENE_MAIN/last_open_file_location'
    _PRODIGAL_BINARY_LOCATION_SETTING = 'GENE_MAIN/prodigal_location'
    _LAST_EXCEL_SAVE_LOCATION_SETTING = 'GENE_MAIN/last_excel_location'
    _GENE_TAB_LABEL = 'Genes'
    _TRNA_TAB_LABEL = 'TRNA'

    def __init__(self, parent=None):
        super(GeneMain, self).__init__(parent)

        # WIDGETS ----------------------------------------------------------------------------------
        # central tab widget
        self.tab = QTabWidget()

        # tables
        self.geneTable = QTableWidget()
        self.trnaTable = QTableWidget()

        # status bar
        self.status = self.statusBar()
        self.status.showMessage('Ready')

        # LAYOUT -----------------------------------------------------------------------------------
        self.setCentralWidget(self.tab)
        # (GRyde) Original size appears to be 400, expanding significantly for better user interface
        self.setMinimumWidth(1000)
        self.setMinimumHeight(1000)

        # ACTIONS ----------------------------------------------------------------------------------
        # new query
        self.newFileAction = self.createAction('&New...', self.fileNew, QKeySequence.New,
                                               tip='Create a new query')

        self.openFileAction = self.createAction('&Open...', self.openFile, QKeySequence.Open,
                                                tip='Open a query file')

        self.saveAsAction = self.createAction('Save as...', self.saveAs,
                                              QKeySequence('Ctrl+Shift+S'),
                                              tip='Save gene data')

        self.saveAction = self.createAction('&Save', self.save, QKeySequence.Save,
                                            tip='Save gene data')

        self.settingsAction = self.createAction('Settings', self.settings, None, )

        self.exportExcelAction = self.createAction('Excel', self.exportExcel, None)

        self.exportGenbankAction = self.createAction('Genbank', self.exportGenbank, None)

        # MENUS ------------------------------------------------------------------------------------
        # file menu
        self.fileMenu = self.menuBar().addMenu('&File')
        self.fileMenuActions = (self.newFileAction, self.openFileAction,
                                self.saveAction, self.saveAsAction)
        self.fileMenu.addActions(self.fileMenuActions)
        self.fileMenu.addSeparator()

        # export submenu
        exportSubMenu = self.fileMenu.addMenu('Export as...')
        exportSubMenu.addAction(self.exportExcelAction)
        exportSubMenu.addAction(self.exportGenbankAction)

        self.fileMenu.addActions([self.settingsAction])

        # VARIABLES --------------------------------------------------------------------------------
        self.queryData = QueryData()
        # whether a file is currently opened
        self.fileOpened = False
        # if unsaved changes are present
        self.dirty = False
        # if saving is enabled
        self.saveEnabled = False

        self.genes = []

        # Get Settings, populate defaults if they do not exist
        self.settings = QSettings(QSettings.IniFormat, QSettings.UserScope, APP_NAME, APP_NAME)
        self.checkDefaultSettings()

        self.enableActions()
        # SETTINGS ---------------------------------------------------------------------------------
        self.setWindowTitle(APP_NAME)

        # check for Prodigal binary
        self.checkProdigal()

    # ACTION METHODS -------------------------------------------------------------------------------
    @pyqtSlot()
    def fileNew(self):
        """
        Action performed when user clicks new file
        """
        # check for unsaved data
        if not self.okToContinue():
            return

        # temporary data in case user quits the dialogs
        tmpQueryData = QueryData()
        # open query dialog
        dialog = NewFileDialog(tmpQueryData, self.settings, self.settings.value(self._PRODIGAL_BINARY_LOCATION_SETTING))
        # if user initiates a query
        if dialog.exec_():
            # query tools
            self.queryData = tmpQueryData
            queryDialog = QueryDialog(self.queryData, self.settings)

            # query to tools is successful
            if queryDialog.exec_():
                # update open variable
                self.fileOpened = True
                self.dirty = True
                self.saveEnabled = False
                # enable / disable actions
                self.enableActions()
                # update window title with temporary file name
                self.setWindowTitle('{} - {}'.format(APP_NAME, 'untitled*'))
                # display gene data
                self.updateTable()
            # query was canceled by user - back to main window
            else:
                pass

        # user does not initiate query - back to main window
        else:
            pass

    def openFile(self):
        """
        Open a query data file
        :param fileName: name of the file
        """
        # check for unsaved progressed
        if not self.okToContinue():
            return

        # open file dialog
        openFileDir = self.settings.value(self._LAST_OPEN_FILE_LOCATION_SETTING)
        fileExtensions = ['GQ Files (*.gq)', 'All Files (*.*)']
        openFileName = QFileDialog.getOpenFileName(self,
                                                   'Open Query File...',
                                                   openFileDir,
                                                   ';;'.join(fileExtensions))

        # check if user provided a file
        if openFileName[0] != '':
            # try to open file
            try:
                with open(openFileName[0], 'rb') as openFile:
                    tempQueryData = pickle.load(openFile)
                    # check if file is QueryData object
                    if not isinstance(tempQueryData, QueryData):
                        # show error message
                        QMessageBox.Warning(self, "Invalid File",
                                            "File: {} was not .gq formatted file".format(
                                                openFileName[0]))
                        return
                    # assign new data
                    self.queryData = tempQueryData
                    self.fileOpened = True
                    self.saveEnabled = True
                    self.enableActions()
                    # save location
                    self.settings.setValue(self._LAST_OPEN_FILE_LOCATION_SETTING, os.path.split(openFileName[0])[0])
                    # change window titles
                    self.setWindowTitle('{} - {}'.format(APP_NAME, openFileName[0]))
                    # update table
                    self.updateTable()

            # opening file was unsuccessful
            except FileNotFoundError:
                QMessageBox.warning(self, 'File Does not Exist',
                                    'File: {} does not exist.'.format(openFileName[0]))
            except Exception as e:
                QMessageBox.warning(self, 'Error Opening File',
                                    str(e))

    @pyqtSlot()
    def save(self):
        """
        Action performed when user clicks save
        Saves changes to file
        :return True/False if save was successful
        """
        # save file
        with open(self.queryData.fileName, 'wb') as saveFile:
            pickle.dump(self.queryData, saveFile)
        # update status bar
        self.status.showMessage('Changes saved to: {}'.format(self.queryData.fileName, 5000))
        return True

    @pyqtSlot()
    def saveAs(self):
        """
        Action performed when user clicks Save As...
        Prompts user for a file name and saves content to file
        :return True/False if save was successful
        """
        # ask user what to save file as
        fileExtensions = ['GQ Files (*.gq)', 'All Files (*.*)']
        saveFileName = QFileDialog.getSaveFileName(self,
                                                   'Save as...',
                                                   '',
                                                   ';;'.join(fileExtensions))

        # check if user didn't provide file
        if saveFileName[0] != '':
            with open(saveFileName[0], 'wb') as saveFile:
                self.queryData.fileName = saveFileName[0]
                pickle.dump(self.queryData, saveFile)
            self.status.showMessage('File saved to: {}'.format(saveFileName[0]), 5000)
            # update file name
            # update window title
            baseFileName = os.path.split(self.queryData.fileName)[1]
            self.setWindowTitle('{} - {}'.format(APP_NAME, baseFileName))
            # allow normal saves
            # self.saveAction.setEnabled(True)
            self.saveEnabled = True
            self.dirty = False
            self.enableActions()
            return True

        return False

    @pyqtSlot()
    def settings(self):
        preferencesDialog = SettingsDialog()
        preferencesDialog.exec_()
        self.updateTable()

    @pyqtSlot()
    def exportExcel(self):
        """
        Save the current table output to a .xlsx file
        """
        # get last saved excel location
        excelLocation = pathlib.Path(self.settings.value(self._LAST_EXCEL_SAVE_LOCATION_SETTING))

        # open a save file dialog
        fileExtensions = ['Excel Spreadsheet (*.xlsx)',
                          'All Files (*.*)']
        excelFileName = QFileDialog.getSaveFileName(self,
                                                    'Save Excel Spreadsheet As...',
                                                    str(excelLocation),
                                                    ';;'.join(fileExtensions))

        # if file name was provided, write to file
        if excelFileName[0] != '':

            GENES_USED = False
            TRNA_USED = False
            for key in self.queryData.toolData.keys():
                if key in GENE_TOOLS:
                    GENES_USED = True
                elif key in TRNA_TOOLS:
                    TRNA_USED = True

            wb = Workbook()
            if GENES_USED:
                self._exportTableToExcel(self.geneTable, 'Genes', wb)
            if TRNA_USED:
                self._exportTableToExcel(self.trnaTable, 'TRNA', wb)

            wb.save(filename=excelFileName[0])

            print(excelFileName[0])
            excelLocation = str(pathlib.Path(excelFileName[0]).parent)
            self.settings.setValue(self._LAST_EXCEL_SAVE_LOCATION_SETTING, excelLocation)

            self.status.showMessage('Exported Excel file to: {}'.format(excelFileName[0]), 5000)

    def _exportTableToExcel(self, table: QTableWidget, label: str, wb: Workbook):
        """
        Adds the given table to the Excel Workbook as a new sheet
        :param table: QTableWidget
        :param label: Name of the new sheet
        :param wb: Excel Workbook
        """

        # check if adding to existing workbook
        # if not, rename first sheet
        if wb.sheetnames[0] == 'Sheet':
            ws = wb['Sheet']
            ws.title = label
        else:
            ws = wb.create_sheet(label)

        # add content to spreadsheet
        # add headers
        currentRow = 1
        for column in range(table.columnCount()):
            headerValue = table.horizontalHeaderItem(column).text()
            cell = ws.cell(row=currentRow, column=column + 1, value=headerValue)
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True)

        # add content
        for row in range(table.rowCount()):
            for column in range(table.columnCount()):
                currCell = table.item(row, column)
                cellValue = currCell.text() if currCell is not None else ''
                if currCell is not None:
                    cellColor = currCell.background().color().getRgb()
                    cellRgbString = ''.join(['{:02x}'.format(num) for num in cellColor[:3]])
                    fontColor = currCell.foreground().color().getRgb()
                    fontRgbString = ''.join(['{:02x}'.format(num) for num in fontColor[:3]])

                # convert an integer string to an integer for spreadsheet functionality
                cellValue = int(cellValue) if cellValue.isdecimal() else cellValue
                cell = ws.cell(row=row + 2, column=column + 1, value=cellValue)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(fgColor=cellRgbString, fill_type='solid')
                cell.font = Font(color=fontRgbString)

    @pyqtSlot()
    def exportGenbank(self):

        # create export dialog
        exportDig = exportGenbankDialog(self.queryData, self.settings)
        if exportDig.exec_():
            # display save status
            self.status.showMessage('Exported Genbank file to: {}'.format(exportDig.saveFileName), 5000)

    # WINDOW METHODS -------------------------------------------------------------------------------

    def closeEvent(self, event):
        if self.okToContinue():
            # exit
            pass
        else:
            event.ignore()

    # HELPER METHODS -------------------------------------------------------------------------------
    def okToContinue(self):
        """
        Checks if any unsaved changes exist and prompts user if there are if they'd like to continue
        Called when user tries to close window
        :return True / False
        """
        if self.dirty:
            userReply = QMessageBox.question(self,
                                             '{} - Unsaved Changes'.format(APP_NAME),
                                             'Save changes? - Data may be lost',
                                             QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel)
            # User cancels - return to main window
            if userReply == QMessageBox.Cancel:
                return False
            # User discards changes - exit
            elif userReply == QMessageBox.No:
                return True
            # User wants to save - save appropriately
            elif userReply == QMessageBox.Yes:
                # if no current save file, prompt for save
                if self.saveAction.isEnabled():
                    self.save()
                else:
                    saveSuccess = self.saveAs()
                    # check if user completed save
                    # don't exit if not saved
                    if not saveSuccess:
                        return False

        return True

    def updateTable(self):
        """
        Displays Gene data to Table
        :return:
        """
        # render tables if data exists
        GENES_COMPLETE = False
        TRNA_COMPLETE = False
        for key in self.queryData.toolData.keys():
            if key in GENE_TOOLS and not GENES_COMPLETE:
                self._update_table(self.geneTable, GENE_TOOLS, 0, self._GENE_TAB_LABEL)
                GENES_COMPLETE = True
            elif key in TRNA_TOOLS and not TRNA_COMPLETE:
                self._update_table(self.trnaTable, TRNA_TOOLS, 1, self._TRNA_TAB_LABEL)
                TRNA_COMPLETE = True

    def _update_table(self, table: QTableWidget, toolList: List[str], index: int, label: str):

        self.tab.removeTab(index)

        # remove any existing cells
        table.setRowCount(0)
        table.setColumnCount(0)

        # table options
        table.setSelectionMode(QTableWidget.NoSelection)

        # sort genes
        genes = []
        usedGeneTools = []
        for tool in self.queryData.toolData:
            if tool in toolList:
                usedGeneTools.append(tool)
                genes += self.queryData.toolData[tool]

        # calculate columns for tools
        toolNumber = len(usedGeneTools)
        toolColumns = toolNumber * 4 + toolNumber - 1
        # add 3 columns for statistics
        totalColumns = toolColumns + 3
        table.setColumnCount(totalColumns)

        # generate headers
        headerIndexes = dict()
        TOTAL_CALLS_COLUMN = 0
        ALL_COLUMN = 1
        ONE_COLUMN = 2
        currIndex = 3
        headers = ['TOTAL CALLS', 'ALL', 'ONE']
        for ind, tool in enumerate(usedGeneTools):
            headerIndexes[tool] = currIndex
            for i in range(4):
                currIndex += 1
                headers.append(tool.upper())
            if ind != toolNumber - 1:
                headers.append('')
                currIndex += 1

        # set headers
        table.setHorizontalHeaderLabels(headers)

        # nothing to display - exit
        if len(genes) == 0:
            # create an empty table
            self.tab.insertTab(index, table, label)
            return

        genes = Gene.GeneUtils.sortGenes(genes)

        # reset genes
        self.genes = []

        # populate table
        # insert first gene
        currentRow = 0
        table.insertRow(currentRow)
        previousGene = genes[0]
        currentGeneCount = 1
        currentGeneSet = [genes[0]]
        currentGenes = dict()
        geneIndex = headerIndexes[previousGene.identity]
        # direction
        directionItem = QTableWidgetItem(previousGene.direction)
        directionItem.setTextAlignment(Qt.AlignCenter)
        table.setItem(currentRow, geneIndex, directionItem)
        # start
        startItem = QTableWidgetItem(str(previousGene.start))
        startItem.setTextAlignment(Qt.AlignCenter)
        table.setItem(currentRow, geneIndex + 1, startItem)
        # stop
        stopItem = QTableWidgetItem(str(previousGene.stop))
        stopItem.setTextAlignment(Qt.AlignCenter)
        table.setItem(currentRow, geneIndex + 2, stopItem)
        # length
        lengthItem = QTableWidgetItem(str(previousGene.length))
        lengthItem.setTextAlignment(Qt.AlignCenter)
        table.setItem(currentRow, geneIndex + 3, lengthItem)

        if previousGene.direction == '+':
            comparingNum = previousGene.start
        else:
            comparingNum = previousGene.stop
        # insert comparing num into dict
        if comparingNum not in currentGenes.keys():
            currentGenes[comparingNum] = 1
        else:
            currentGenes[comparingNum] += 1

        # insert rest of genes
        for gene in genes[1:]:
            geneIndex = headerIndexes[gene.identity]

            # same gene - add to current row
            if gene == previousGene:
                currentGeneCount += 1
                currentGeneSet.append(gene)
            # different gene - create new row
            else:
                # record TOTAL_CALLS, ALL and ONE for previous gene
                totalItem = QTableWidgetItem(str(currentGeneCount))
                totalItem.setTextAlignment(Qt.AlignCenter)
                table.setItem(currentRow, TOTAL_CALLS_COLUMN, totalItem)
                if currentGeneCount == toolNumber:
                    allItem = QTableWidgetItem(str('X'))
                    allItem.setTextAlignment(Qt.AlignCenter)
                    table.setItem(currentRow, ALL_COLUMN, allItem)
                elif currentGeneCount == 1:
                    oneItem = QTableWidgetItem(str('X'))
                    oneItem.setTextAlignment(Qt.AlignCenter)
                    table.setItem(currentRow, ONE_COLUMN, oneItem)

                # color row
                colorSetting = self.settings.value(ColorTable.CELL_COLOR_SETTING + str(currentGeneCount - 1))
                colorNums = [int(num) for num in colorSetting.split(' ')]
                color = QColor(*colorNums)
                # color text
                textColorSetting = self.settings.value(ColorTable.MAJORITY_TEXT_SETTING + str(currentGeneCount - 1))
                textNums = [int(num) for num in textColorSetting.split(' ')]
                textColor = QColor(*textNums)
                # (GRyde) ******************************************************************** 
                # Establish 2nd/3rd/4th majority text settings
                # 2nd
                textColorSecondSetting = self.settings.value(ColorTable.MAJORITY_TEXT_SECOND_SETTING + str(currentGeneCount - 1))
                textNumsSecond = [int(num) for num in textColorSecondSetting.split(' ')]
                textColorSecond = QColor(*textNumsSecond)
                # 3rd
                textColorThirdSetting = self.settings.value(ColorTable.MAJORITY_TEXT_THIRD_SETTING + str(currentGeneCount - 1))
                textNumsThird = [int(num) for num in textColorThirdSetting.split(' ')]
                textColorThird = QColor(*textNumsThird)
                # 4th
                textColorFourthSetting = self.settings.value(ColorTable.MAJORITY_TEXT_FOURTH_SETTING + str(currentGeneCount - 1))
                textNumsFourth = [int(num) for num in textColorFourthSetting.split(' ')]
                textColorFourth = QColor(*textNumsFourth)
                # 5th
                textColorFifthSetting = self.settings.value(ColorTable.MAJORITY_TEXT_FIFTH_SETTING + str(currentGeneCount - 1))
                textNumsFifth = [int(num) for num in textColorFifthSetting.split(' ')]
                textColorFifth = QColor(*textNumsFifth)
                # 6th
                textColorSixthSetting = self.settings.value(ColorTable.MAJORITY_TEXT_SIXTH_SETTING + str(currentGeneCount - 1))
                textNumsSixth = [int(num) for num in textColorSixthSetting.split(' ')]
                textColorSixth = QColor(*textNumsSixth)
                # 7th
                textColorSeventhSetting = self.settings.value(ColorTable.MAJORITY_TEXT_SEVENTH_SETTING + str(currentGeneCount - 1))
                textNumsSeventh = [int(num) for num in textColorSeventhSetting.split(' ')]
                textColorSeventh = QColor(*textNumsSeventh)
                # 8th
                textColorEighthSetting = self.settings.value(ColorTable.MAJORITY_TEXT_EIGHTH_SETTING + str(currentGeneCount - 1))
                textNumsEighth = [int(num) for num in textColorEighthSetting.split(' ')]
                textColorEighth = QColor(*textNumsEighth)
                # Original start
                # Establish minority text settings
                minTextColorSetting = self.settings.value(ColorTable.MINORITY_TEXT_SETTING + str(currentGeneCount - 1))
                minTextNums = [int(num) for num in minTextColorSetting.split(' ')]
                minTextColor = QColor(*minTextNums)
                # Original end
                # (GRyde) ******************************************************************** end
                # TODO: Figure out minority rule
                # (GRyde) ******************************************************************** start
                
                #Establish if majority case should be compared against start or stop
                if currentGeneSet[0].direction == '+':
                    compareStart = True
                else:
                    compareStart = False
                    
                majorityNum = []
                majorityCalls = 0
                
                uniqueNum = []
                
                # Determine start/stop with greatest amount of calls for current gene
                for number in currentGenes.keys():
                    if currentGenes[number] > majorityCalls:
                        majorityCalls = currentGenes[number]
                        
                # Compare calls to each comparingNum vs highest amount for current gene
                for number in currentGenes.keys():
                    if currentGenes[number] == majorityCalls:
                        majorityNum.append(number)
                        
                # Including minTextColor just for testing, remove once other 5 colors added        
                textColors = [
                    textColor,
                    textColorSecond,
                    textColorThird,
                    textColorFourth,
                    textColorFifth,
                    textColorSixth,
                    textColorSeventh,
                    textColorEighth,
                    minTextColor
                ]        
                textList = []
                
                # If all start/stop comparison numbers match, no need to set up different colors
                if not (majorityCalls == currentGeneCount):
                    
                    for number in majorityNum:
                        uniqueNum.append(number)
                        
                    for number in currentGenes.keys():
                        if not (number in uniqueNum):
                            uniqueNum.append(number)
                            
                    for currentGene in currentGeneSet:
                        if compareStart:
                            colorIndex = uniqueNum.index(currentGene.start)
                            textList.append([currentGene, textColors[colorIndex]])
                        else:
                            colorIndex = uniqueNum.index(currentGene.stop)
                            textList.append([currentGene, textColors[colorIndex]])
                
                # Ver. 2 start
                # if not (majorityCalls == currentGeneCount):
                    # for currentGene in currentGeneSet:
                        # matchFound = False
                        # for number in majorityNum:
                            # if compareStart:
                                # if currentGene.start == number:
                                    # colorIndex = majorityNum.index(number)
                                    # textList.append([currentGene, textColors[colorIndex]])
                                    # matchFound = True
                                    # break
                            # else:
                                # if currentGene.stop == number:
                                    # colorIndex = majorityNum.index(number)
                                    # textList.append([currentGene, textColors[colorIndex]])
                                    # matchFound = True
                                    # break
                        # if not matchFound:
                            # textList.append([currentGene, minTextColor])
                # Ver. 2 end
                
                # Original start
                # If single comparing number received all tool calls, ignore as there is no minority
                # Otherwise, compare start/stop of each gene in currentGeneSet vs majority numbers
                # if not (majorityCalls == currentGeneCount):
                    # for minGene in currentGeneSet:
                        # isMinority = True
                        # for number in majorityNum:
                            # if compareStart:
                                # if minGene.start == number:
                                    # isMinority = False
                            # else:
                                # if minGene.stop == number:
                                    # isMinority = False
                        # minTextList.append([isMinority, minGene])
                # Original end
                                     
                # (GRyde) ******************************************************************** end
                for column in range(totalColumns):
                    item = table.item(currentRow, column)
                    # insert blank item if none is present
                    if item is None:
                        item = QTableWidgetItem('')
                        table.setItem(currentRow, column, item)
                    item.setBackground(color)
                    item.setForeground(textColor)
                    
                # (GRyde) ******************************************************************** start
                for currentGene, geneColor in textList:
                    geneColIndex = headerIndexes[currentGene.identity]
                    if compareStart:
                        item = table.item(currentRow, geneColIndex + 1)
                        item.setForeground(geneColor)
                    else:
                        item = table.item(currentRow, geneColIndex + 2)
                        item.setForeground(geneColor)
                # Original start
                # Set minority text for appropriate items
                # for minCheck, minGene in minTextList:
                    # if minCheck:
                        # minGeneIndex = headerIndexes[minGene.identity]
                        # if compareStart:
                            # item = table.item(currentRow, minGeneIndex + 1)
                            # item.setForeground(minTextColor)
                        # else:
                            # item = table.item(currentRow, minGeneIndex + 2)
                            # item.setForeground(minTextColor)
                # Original end
                # (GRyde) ******************************************************************** end

                # record new gene
                currentGeneCount = 1
                currentGenes = dict()
                currentRow += 1
                table.insertRow(currentRow)

                # add full set to genes
                self.genes.append(currentGeneSet)
                currentGeneSet = [gene]

            # add gene to table
            # direction
            directionItem = QTableWidgetItem(gene.direction)
            directionItem.setTextAlignment(Qt.AlignCenter)
            table.setItem(currentRow, geneIndex, directionItem)
            # start
            startItem = QTableWidgetItem(str(gene.start))
            startItem.setTextAlignment(Qt.AlignCenter)
            table.setItem(currentRow, geneIndex + 1, startItem)
            # stop
            stopItem = QTableWidgetItem(str(gene.stop))
            stopItem.setTextAlignment(Qt.AlignCenter)
            table.setItem(currentRow, geneIndex + 2, stopItem)
            # length
            lengthItem = QTableWidgetItem(str(gene.length))
            lengthItem.setTextAlignment(Qt.AlignCenter)
            table.setItem(currentRow, geneIndex + 3, lengthItem)

            if gene.direction == '+':
                comparingNum = gene.start
            else:
                comparingNum = gene.stop
                # insert comparing num into dict
            if comparingNum not in currentGenes.keys():
                currentGenes[comparingNum] = 1
            else:
                currentGenes[comparingNum] += 1

            # set gene to compare next against
            previousGene = gene

        # record TOTAL_CALLS, ALL and ONE for last gene
        totalItem = QTableWidgetItem(str(currentGeneCount))
        totalItem.setTextAlignment(Qt.AlignCenter)
        table.setItem(currentRow, TOTAL_CALLS_COLUMN, totalItem)
        if currentGeneCount == toolNumber:
            allItem = QTableWidgetItem(str('X'))
            allItem.setTextAlignment(Qt.AlignCenter)
            table.setItem(currentRow, ALL_COLUMN, allItem)
        elif currentGeneCount == 1:
            oneItem = QTableWidgetItem(str('X'))
            oneItem.setTextAlignment(Qt.AlignCenter)
            table.setItem(currentRow, ONE_COLUMN, oneItem)

        # append last set of genes
        self.genes.append(currentGeneSet)

        # color last row
        colorSetting = self.settings.value(ColorTable.CELL_COLOR_SETTING + str(currentGeneCount - 1))
        colorNums = [int(num) for num in colorSetting.split(' ')]
        color = QColor(*colorNums)
        # color text
        textColorSetting = self.settings.value(ColorTable.MAJORITY_TEXT_SETTING + str(currentGeneCount - 1))
        textNums = [int(num) for num in textColorSetting.split(' ')]
        textColor = QColor(*textNums)
        # TODO: Figure out minority rule
        for column in range(totalColumns):
            item = table.item(currentRow, column)
            # insert blank item if blank cell
            if item is None:
                item = QTableWidgetItem('')
                table.setItem(currentRow, column, item)
            item.setBackground(color)
            item.setForeground(textColor)
            
        # table font
        tableFont = QFont()
        tableFont.setPointSize(15)
        
        for row in range(currentRow + 1):
            for column in range(totalColumns):
                item = table.item(row, column)
                item.setFont(tableFont)

        # show tab
        # self.tab.addTab(table, self._GENE_TAB_LABEL)
        self.tab.insertTab(index, table, label)

    def enableActions(self):
        """
        Enables / Disables GUI actions
        :return:
        """
        # file open actions
        if self.fileOpened:
            self.saveAsAction.setEnabled(True)
            self.exportExcelAction.setEnabled(True)
            self.exportGenbankAction.setEnabled(True)
        else:
            self.saveAsAction.setEnabled(False)
            self.exportExcelAction.setEnabled(False)
            self.exportGenbankAction.setEnabled(False)

        if self.saveEnabled:
            self.saveAction.setEnabled(True)
        else:
            self.saveAction.setEnabled(False)

    def checkProdigal(self):

        prodigalPath = self.settings.value(self._PRODIGAL_BINARY_LOCATION_SETTING)
        # if binary does not exist or binary has disappeared, prompt to download
        if prodigalPath is None or not os.path.exists(prodigalPath):
            currRelease = ProdigalRelease()
            # prompt to download prodigal
            # location to store binary is gquery's folder
            td = ThreadData(pathlib.Path(__file__).parent)
            prodigalDownloadDig = phagecommander.GuiWidgets.ProdigalDownloadDialog(currRelease, td)
            if prodigalDownloadDig.exec_():
                self.settings.setValue(self._PRODIGAL_BINARY_LOCATION_SETTING, td.data)
            else:
                self.settings.setValue(self._PRODIGAL_BINARY_LOCATION_SETTING, None)

    def createAction(self, text, slot=None, shortcut=None, icon=None, tip=None, checkable=False,
                     signal='triggered()'):
        """
        Helper method for generating QActions
        :param text: name of the action
        :param slot: pyqtSlot
        :param shortcut: shortcut to map to
        :param icon: path to action icon
        :param tip: tooltip help string
        :param checkable: True if an ON/OFF action
        :param signal:
        :return: QAction
        """
        action = QAction(text, self)
        if slot is not None:
            action.triggered.connect(slot)
        if shortcut is not None:
            action.setShortcut(shortcut)
        if icon is not None:
            action.setIcon(QIcon(icon))
        if tip is not None:
            action.setToolTip(tip)
            action.setStatusTip(tip)
        if checkable:
            action.setCheckable(True)

        return action

    def checkDefaultSettings(self):
        """
        Check for require default settings
        If they do not exist - populate them
        """
        # NewFileDialog SETTINGS
        NewFileDialog.checkDefaultSettings(self.settings)
        # EXPORT GENBANK SETTINGS
        exportGenbankDialog.checkDefaultSettings(self.settings)
        # COLOR SETTINGS
        ColorTable.checkDefaultSettings(self.settings)

        # OPEN FILE LOCATION
        if self.settings.value(self._LAST_OPEN_FILE_LOCATION_SETTING) is None:
            self.settings.setValue(self._LAST_OPEN_FILE_LOCATION_SETTING, '')

        # EXCEL SAVE LOCATION
        if self.settings.value(self._LAST_EXCEL_SAVE_LOCATION_SETTING) is None:
            self.settings.setValue(self._LAST_EXCEL_SAVE_LOCATION_SETTING, '')


# MAIN FUNCTION
def main():
    app = QApplication([])
    window = GeneMain()
    window.show()
    app.exec_()


if __name__ == '__main__':
    main()
